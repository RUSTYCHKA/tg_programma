import uuid
import ssl
import dill as pickle
import openpyxl
from itertools import cycle
from datetime import datetime, timedelta, timezone
from asyncio.exceptions import IncompleteReadError
import calendar
from langdetect import detect
import json
import os
import asyncio
import string
import logging
import random
import re
from random import randint as rd
from openai import OpenAI
from sqlite3 import OperationalError
from telethon import TelegramClient, errors, functions, types, events
from telethon.tl.functions.channels import JoinChannelRequest
from telethon.tl.functions.messages import ImportChatInviteRequest, GetAllStickersRequest, GetStickerSetRequest
from telethon.tl.types import (
    UserStatusEmpty,
    UserStatusLastMonth,
    UserStatusLastWeek,
    UserStatusOffline,
    UserStatusOnline,
    UserStatusRecently,
    InputPhoto,
    InputChatPhoto,
    ReactionEmoji,
    ChannelParticipantsBanned,
    MediaAreaUrl,
    MediaAreaCoordinates,
    InputStickerSetID,
    InputChannelEmpty
)
from telethon.errors import UserAlreadyParticipantError, PeerIdInvalidError, FloodWaitError, MsgIdInvalidError, PeerFloodError
from telethon.tl.functions.account import UpdateProfileRequest, UpdateUsernameRequest
from telethon.tl.functions.photos import DeletePhotosRequest, UploadProfilePhotoRequest, GetUserPhotosRequest
from telethon.tl.functions.stories import ReadStoriesRequest, SendReactionRequest, GetPeerStoriesRequest
from telethon.tl.functions.users import GetFullUserRequest
from telethon.errors import UsernameInvalidError, ChannelPrivateError
from config import OPEN_AI_TOKEN
from core.functions import GetBio, ProxyFromUrl, edit_message_time, generate_g_a_hash, message_handler, message_handler_wiretapping, post_handler, prepare_answer, promt_generation, reaction1, reaction2, reaction3
from core.progress import ProgressBar
from core.texts import STARTED_TASK, progress, TASK_NAMES
from core import kb
from io import StringIO, BytesIO
from handlers.start import bot
from pol.predict import predict
from pol.predict import model
from webapp_ import active_tasks

report_reasons = [
    types.InputReportReasonChildAbuse(),
    types.InputReportReasonCopyright(),
    types.InputReportReasonFake(),
    types.InputReportReasonGeoIrrelevant(),
    types.InputReportReasonIllegalDrugs(),
    types.InputReportReasonOther(),
    types.InputReportReasonPersonalDetails(),
    types.InputReportReasonPornography(),
    types.InputReportReasonSpam(),
    types.InputReportReasonViolence()
]


def remove_timezone(dt):
    """Убирает временную зону, если объект — это datetime"""
    if isinstance(dt, datetime) and dt.tzinfo is not None:
        return dt.replace(tzinfo=None)
    return dt


def create_excel_with_multiple_sheets(result_users, bio):
    # Создаем новый Excel файл в памяти

    output = BytesIO()
    workbook = openpyxl.Workbook()

    # Удаляем дефолтный пустой лист
    workbook.remove(workbook.active)

    # Проходим по каждому чату (ключ в словаре result_users)
    for chat_id, users in result_users.items():
        # Создаем новый лист для каждого чата
        chat_id = chat_id.replace("https://t.me/", "")
        worksheet = workbook.create_sheet(title=f"{chat_id}")

        # Записываем заголовки колонок
        if bio == True:
            worksheet.append(["user_id", "username", "phone",
                              "status", "premium", "story", "pol", "bio"])
        else:
            worksheet.append(["user_id", "username", "phone",
                              "status", "premium", "story", "pol"])
        column_widths = [15, 20, 15, 20, 15, 10, 10,
            200] if bio else [15, 20, 15, 20, 15, 10, 10]
        for i, width in enumerate(column_widths):
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(
                i + 1)].width = width
        # Заполняем данные для каждого пользователя в текущем чате
        for user_id, data in users.items():
            worksheet.append([
                user_id,
                f"@{data['username']}" if data['username'] else "",
                data['phone'],
                remove_timezone(data['status']),
                data['premium'],
                data['story'],
                data['pol'],
                data['bio']
            ])

    # Сохраняем данные в BytesIO
    workbook.save(output)
    output.seek(0)

    # Возвращаем файл в байтах
    return output.getvalue()


def distribute_users(users, accounts, users_per_account):
    distribution = {account: [] for account in accounts}
    account_iterator = cycle(accounts)

    for user in users:
        account = next(account_iterator)
        while len(distribution[account]) >= users_per_account:
            account = next(account_iterator)
        distribution[account].append(user)

    return distribution


async def update_2fa(client, password):

    with open(f'{client.replace(".session", ".json")}', 'r') as f:
        data = json.load(f)

    data['twoFA'] = password

    with open(f'{client.replace(".session", ".json")}', 'w') as f:
        json.dump(data, f, indent=4, ensure_ascii=False)


def get_relative_time(expires_at):
    now = datetime.now(timezone.utc)

    expires_at = expires_at.astimezone(timezone.utc).replace(tzinfo=None)
    now = now.replace(tzinfo=None)

    delta = now - expires_at
    seconds = int(round(delta.total_seconds(), 0))
    if delta < timedelta(seconds=0):
        return "Недавно", ""

    if delta < timedelta(seconds=60):
        return f"{int(delta.total_seconds())} секунд назад", seconds
    elif delta < timedelta(minutes=60):
        return f"{int(delta.total_seconds() // 60)} минут назад", seconds
    elif delta < timedelta(hours=24):
        return f"{int(delta.total_seconds() // 3600)} часов назад", seconds
    elif delta < timedelta(days=7):
        return f"{int(delta.days)} дней назад", seconds
    elif delta < timedelta(days=30):
        return f"{int(delta.days)} дней назад", seconds
    elif delta < timedelta(weeks=4):
        return f"{int(delta.days // 7)} недель назад", seconds
    elif delta < timedelta(days=365):
        return f"{calendar.month_name[expires_at.month]} {expires_at.year}", seconds
    else:
        return f"{expires_at.year}", seconds


def chunks(lst, n):
    """Yield successive n-sized chunks from lst."""
    for i in range(0, len(lst), n):
        yield lst[i:i + n]


def random_string(length):
    return ''.join(random.sample(string.ascii_letters + string.digits, length))


def group_list(lst, n):
    return [lst[i:i+n] for i in range(0, len(lst), n)]


class Task:
    def __init__(self, work_dir: str, session_names: list, progress_bar: ProgressBar, proxies: list = None, sessions_count: int = None, data=None, logger=None):
        self.sessions = []
        self.work_dir = work_dir
        self.data = data
        self.count_errors = 0
        self.progress_bar = progress_bar
        self.proxies = proxies
        self.sessions_count = sessions_count
        self.tasks = []
        self.tasks_machine = []
        self.remaining = []

        try:
            self.task_id = data["task_id"]
        except:
            self.task_id = None
        try:
            self.id = data["id"]
        except:
            self.id = None
        try:
            self.path = f"tasks/{data["customer"].id}/{data["task_id"]}/{data['id'].split('=')[-1]}"
        except:
            self.path = None
        if self.task_id == "2fa":
            self.path = f"{data["customer"].id}"
        self.stream = StringIO()
        self.datetime = datetime.now()
        self.handler = logging.StreamHandler(self.stream)
        self.logger = logging.getLogger(
            self.path) if self.path is not None else logger

        proxy_index = 0
        for session_name in session_names:
            proxy = proxies[proxy_index] if proxies else None
            self.sessions.append(
                Session(f"{work_dir}/{session_name}.session", proxy, logger=self.logger))
            proxy_index = (proxy_index + 1) % len(proxies) if proxies else 0
        if self.sessions_count is not None:
            self.sessions = self.sessions[:self.sessions_count]
        try:
            if data["customer"].id not in active_tasks:
                active_tasks[data["customer"].id] = [self]
            else:
                active_tasks[data["customer"].id].append(self)
        except:
            pass

    def _status_text(self, status):
        if status is True:
            return 'Работает ✅'
        elif "ограничения" in status:
            return "Спам-блок ⚠"
        else:
            return f'{status} ❌'

    def set_name(self, name):
        self.name = name

    async def Start(self):
        for s in self.sessions:
            self.tasks.append(asyncio.create_task(
                s.Start()))

    async def Checking(self, spam_block: bool = False):
        checked = {}
        total_count = len(self.sessions)
        tasks = [s.Check(spam_block) for s in self.sessions]
        results = await asyncio.gather(*tasks)

        for i, (s, status) in enumerate(zip(self.sessions, results)):
            if self.sessions_count is not None and len(checked) == self.sessions_count:
                return checked

            session_name = s.session.replace(
                ".session", "").replace(f"{self.work_dir}/", "")
            checked[session_name] = status

            if status is not True and "ограничен" not in status:
                to_delete = [s.session]
                try:
                    os.remove(f"{self.work_dir}/{session_name}.session")
                except:
                    pass

                if "Не смог прочитать" not in status:
                    os.remove(f"{self.work_dir}/{session_name}.json")

        return checked

    async def SetMaxFlood(self, max_flood: int):
        for s in self.sessions:
            try:
                s.client.flood_sleep_threshold = max_flood
            except:
                print(s)

    async def ChangeInfo(self, firstnames, lastnames, descriptions, usernames, photos, stories):
        for i, s in enumerate(self.sessions):
            self.tasks.append(asyncio.create_task(s.ChangeInfo(
                firstnames[i], lastnames[i], descriptions[i], usernames[i], photos[i], stories[i])))
        result = []
        while True:
            for i, s in enumerate(self.sessions):
                try:
                    res = self.tasks[i].result()
                    result.append(res)
                except asyncio.exceptions.InvalidStateError:
                    pass
            if len(result) >= len(self.sessions):
                return result
            await asyncio.sleep(1)

    async def GetAccountsInfoIquery(self, spam_block=False):

        async def process_session(s):
            if not s.client.is_connected():
                status = await s.Check(spam_block)
                if status is not True and "ограничен" not in status and "заморожен" not in status:
                    session_name = s.session.replace(
                        ".session", "").replace(f"{self.work_dir}/", "")
                    user_info = {
                        'username': "Аккаунт забанен",
                        'phone': "Аккаунт забанен",
                        'first_name': "Аккаунт забанен",
                        'last_name': "Аккаунт забанен",
                        'status': status,
                        'id': "Аккаунт забанен",
                        'is_premium': "Аккаунт забанен"
                    }
                    return session_name, None, user_info
                elif status is not True and "заморожен" in status:
                    me = await s.client.get_me()
                    session_name = s.session.replace(
                        ".session", "").replace(f"{self.work_dir}/", "")
                    user_info = {
                        'username': me.username,
                        'phone': me.phone,
                        'first_name': "",
                        'last_name': "Аккаунт заморожен",
                        'status': status,
                        'id': me.id,
                        'is_premium': me.premium
                    }
                    return session_name, None, user_info

            session_name = s.session.replace(
                ".session", "").replace(f"{self.work_dir}/", "")
            try:
                me = await s.client.get_me()
            except:
                session_name = s.session.replace(
                    ".session", "").replace(f"{self.work_dir}/", "")
                user_info = {
                    'username': "Аккаунт забанен",
                    'phone': "Аккаунт забанен",
                    'first_name': "Аккаунт забанен",
                    'last_name': "Аккаунт забанен",
                    'status': status,
                    'id': "Аккаунт забанен",
                    'is_premium': "Аккаунт забанен"
                }
                return session_name, None, user_info

            user_info = {
                'username': me.username,
                'phone': me.phone,
                'first_name': me.first_name,
                'last_name': me.last_name,
                'status': status,
                'id': me.id,
                'is_premium': me.premium

            }
            await s.client.disconnect()
            return session_name, None, user_info

        tasks = [process_session(s) for s in self.sessions]
        results = await asyncio.gather(*tasks)

        info = {}
        for session_name, error, user_info in results:
            if error:
                return None, error
            if session_name and user_info:
                info[session_name] = user_info

        return info

    async def GetUsernames(self, chat_id):
        text = []
        self.progress_bar.hint = "Для корректной работы добавьте данные аккаунта в администраторы чата и разрешите добавлять администраторов."
        for i, s in enumerate(self.sessions):
            if not s.client.is_connected():
                status = await s.Check()
                if status is not True:
                    continue

            if not s.me.username:
                while True:
                    username = random_string(8)
                    try:
                        await s.client(functions.account.UpdateUsernameRequest(username))
                    except (errors.rpcerrorlist.UsernameOccupiedError, errors.rpcerrorlist.UsernameNotModifiedError, errors.rpcerrorlist.UsernameInvalidError):
                        continue
                    else:
                        break
                s.me.username = username
            status = await s.Join(chat_id)

            session_hint = f'\t<i>{i+1}. {s.session.replace(".session", "").replace(f"{self.work_dir}/", "")}</i> - '
            if status is True:
                text.append(session_hint + "<code>" +
                            s.me.username + "</code>")
            else:
                await self.CancelAll()
                await self.StopAll()
                await self._logging_task(self.task_id)
                text.append(session_hint + status[1])

            await self.progress_bar.Process(progress.TASK_PROGRESS.format(info="\n".join(text)), reply_markup=kb.start_inviting)
            return i

    async def GetAccountsInfo(self):
        info = {}
        for i, s in enumerate(self.sessions):
            session_name = s.session.replace(
                ".session", "").replace(f"{self.work_dir}/", "")
            if s.me is None:
                continue
            user_info = {'user_id': s.me.id, 'username': s.me.username, 'phone': s.me.phone if s.me else None,
                         'proxy': s.proxy, 'spam_block': "Да" if s.spamblock is True else "Нет"}
            info[session_name] = user_info

        json_data = json.dumps(info, indent=2).encode('utf-8')
        csv_data = "Имя сессии,User_ID,Юзернейм,Телефон,Прокси,Спам-блок\n" + \
            ("\n".join([f"{session_name},{data['user_id']},{data['username']},{data['phone']},{data.get('proxy')},{data['spam_block']}".replace(
                "\n", "") for session_name, data in info.items()]))
        csv_data = csv_data.encode('utf-8')
        return json_data, csv_data

    async def Parsing(self, chats, wait, bio):
        for i, chunk in enumerate(chats):
            try:
                s = self.sessions[i]
            except IndexError:
                break
            self.tasks.append(asyncio.create_task(
                s.Parsing(chunk, 0, 1, wait, bio)))
            self.progress_bar.hint = TASK_NAMES['parsing']
        while True:
            result = await self._logging_task('parsing')
            await self._task_start()

            await asyncio.sleep(100000000)

    async def ParsingSearch(self, identifiers, limit_search):
        for i, chunk in enumerate(identifiers):
            try:
                s = self.sessions[i]
            except IndexError:
                break
            self.tasks.append(asyncio.create_task(
                s.ParsingSearch(chunk, limit_search)))
            self.progress_bar.hint = TASK_NAMES['parsing_search']
        while True:
            result = await self._logging_task('search_chats')
            if result is True:
                result_keywords = {
                    keyword: data for s in self.sessions for keyword, data in s.result.items()}

                return result_keywords
            await asyncio.sleep(1)

    async def Spaming(self, users, wait, text=None, photo=None, video=None, message_id=None, channel_id=None, count=0, remaining=None, via_bot=None, sticker=None):
        self.remaining = remaining
        for i, chunk in enumerate(users):
            try:
                s = self.sessions[i]
            except IndexError:
                break
            self.tasks.append(asyncio.create_task(
                s.Spaming(chunk, wait, text, photo, video, message_id, channel_id, count, remaining, via_bot, sticker)))
            self.progress_bar.hint = TASK_NAMES['spam']

        while True:
            result = await self._logging_task('spaming')
            await self._task_start()
            if result is True:
                break
            await asyncio.sleep(100000)

    async def Tipping(self, urls):
        for i, s in enumerate(self.sessions):
            self.tasks.append(asyncio.create_task(s.Tipping(urls)))
            self.progress_bar.hint = TASK_NAMES['tipping']
        while True:
            result = await self._logging_task('tipping')
            await self._task_start()
            if result is True:
                break
            await asyncio.sleep(100000)

    async def AnswerMachine(self, url):
        for i, s in enumerate(self.sessions):
            self.tasks_machine.append(
                asyncio.create_task(s.AnswerMachine(url)))

    async def FetchMailing(self, streams, delay_join, delay_send, counts_message, count_mailings, delay_mailing, file_id, text, urls, answer_machine, url_answer_machine):
        for i, chunk in enumerate(urls):
            try:
                s = self.sessions[i]
            except IndexError:
                break

            self.tasks.append(asyncio.create_task(s.FetchMailing(streams, delay_join, delay_send,
                              counts_message, count_mailings, delay_mailing, file_id, text, chunk)))
            if answer_machine:
                self.tasks.append(asyncio.create_task(
                    s.AnswerMachine(url_answer_machine)))
        while True:
            self.progress_bar.hint = TASK_NAMES["autoposting"]
            result = await self._logging_task('mailing')
            await self._task_start()
            if result is True:
                break
            await asyncio.sleep(100000)

    async def SendMentions(self, chats, wait, send, text=None, photo=None, video=None):
        for i, chunk in enumerate(chats):
            try:
                s = self.sessions[i]
            except IndexError:
                break
            self.tasks.append(asyncio.create_task(
                s.SendMentions(chunk, wait, send, text, photo, video)))
            self.progress_bar.hint = TASK_NAMES['mentions']
        while True:
            result = await self._logging_task('mentions')
            await self._task_start()
            if result is True:
                break
            await asyncio.sleep(100000)

    async def Inviting(self, users, chat_id, wait, num_session, count=0, remaining=None):
        if len(self.sessions) > 1:
            s_ = self.sessions[num_session]
            # self.sessions.remove(s_)
            for s in self.sessions:
                if not s.client.is_connected():
                    status = await s.Check()
                    if status is not True:
                        continue
                try:
                    status = await s.Join(chat_id)
                except Exception as e:
                    self.logger.error(e)
            try:
                await s_.AllAdmin(self.sessions, chat_id)
            except Exception as e:
                self.logger.error(e)
        for i, chunk in enumerate(users):
            try:
                s = self.sessions[i]
            except IndexError:
                break
            self.tasks.append(asyncio.create_task(
                s.Inviting(chunk, chat_id, wait, count, remaining)))
            self.progress_bar.hint = TASK_NAMES['inviting']
        while True:
            result = await self._logging_task('inviting')
            await self._task_start()
            if result is True:
                break
            await asyncio.sleep(100000)

    async def MassLooking(self, reactions, looking, reaction_flood, stories_account, stories_user, identifiers, reaction_count):
        for i, chunk in enumerate(identifiers):
            try:
                s = self.sessions[i]
            except IndexError:
                break

            reaction = random.choice(reactions)
            self.tasks.append(asyncio.create_task(
                s.MassLooking(reaction, looking, reaction_flood, stories_account, stories_user, chunk, reaction_count)))
            self.progress_bar.hint = TASK_NAMES['masslooking']
        while True:

            result = await self._logging_task('masslooking')
            await self._task_start()
            if result is True:
                break
            await asyncio.sleep(10000000000)

    async def MassLookingChats(self, reactions, chats, safe_mode):
        for i, chunk in enumerate(chats):
            try:
                s = self.sessions[i]
            except IndexError:
                break
            self.tasks.append(s.client.loop.create_task(
                s.MassLookingChats(reactions, chunk, safe_mode)))
            self.progress_bar.hint = TASK_NAMES['masslooking']
        while True:
            result = await self._logging_task('masslookingchats')
            await self._task_start()
            if result is True:
                break
            await asyncio.sleep(1000000)

    async def MassTeging(self, wait, users, stories_url, photo, time, videos=None):
        for i, s in enumerate(self.sessions):
            try:

                self.tasks.append(asyncio.create_task(
                    s.MassTeging(wait, users[i], stories_url, photo[i], time, videos[i])))
            except Exception as e:
                print(e)
            self.progress_bar.hint = TASK_NAMES['massteging']
        while True:
            result = await self._logging_task('massteging')
            await self._task_start()
            if result is True:
                break
            await asyncio.sleep(100000)

    async def NumberCheck(self, wait, numbers):
        for i, s in enumerate(self.sessions):
            try:
                self.tasks.append(asyncio.create_task(
                    s.NumberCheck(wait, numbers[i])))
            except Exception as e:
                print(e)
            self.progress_bar.hint = TASK_NAMES['number_check']
        while True:
            result = await self._logging_task('number_check')
            await self._task_start()
            if result is True:
                break
            await asyncio.sleep(100000)

    async def Neurocommenting(self, channels, promt, url_answer_machine, type, flood, api):

        for i, chunk in enumerate(channels):
            try:
                s = self.sessions[i]
            except IndexError:
                break
            self.tasks.append(asyncio.create_task(
                s.Neurocommenting(chunk, promt, url_answer_machine, type, flood, api)))

            self.progress_bar.hint = TASK_NAMES['neurocommenting']
        while True:
            result = await self._logging_task('neurocommenting')
            await self._task_start()
            if result is True:
                break
            await asyncio.sleep(100000)

    async def Commenting(self, channels, wait, type, text=None, photo=None, video=None, message_id=None, channel_id=None,  via_bot=None, sticker=None):
        for i, s in enumerate(self.sessions):
            try:
                self.tasks.append(asyncio.create_task(
                    s.Commenting(channels[i], wait, type, text, photo, video, message_id, channel_id, via_bot, sticker)))
            except:
                pass
            self.progress_bar.hint = TASK_NAMES['commenting']
        while True:
            result = await self._logging_task('commenting')
            await self._task_start()
            if result is True:
                break
            await asyncio.sleep(100000)

    async def CommentingEdit(self, channels, wait, entry, type, first_comment, text=None, photo=None, video=None, message_id=None, channel_id=None,  via_bot=None, sticker=None):
        for i, s in enumerate(self.sessions):
            try:
                self.tasks.append(asyncio.create_task(
                    s.CommentingEdit(channels[i], wait, entry, type, first_comment, text, photo, video, message_id, channel_id, via_bot, sticker)))
            except:
                pass
            self.progress_bar.hint = TASK_NAMES['commenting']
        while True:
            result = await self._logging_task('commenting')
            await self._task_start()
            if result is True:
                break
            await asyncio.sleep(100000)

    async def Reports(self, wait, chats, report_text, report_reason):
        for i, s in enumerate(self.sessions):
            self.tasks.append(asyncio.create_task(
                s.Reports(wait, *chats, report_text, report_reason)))
            self.progress_bar.hint = TASK_NAMES['reports']
        while True:
            result = await self._logging_task('reports')
            await self._task_start()
            if result is True:
                break
            await asyncio.sleep(100000)

    async def ResetAuthorization(self):
        for i, s in enumerate(self.sessions):
            self.tasks.append(asyncio.create_task(s.ResetAuthorization()))
        result = []
        while True:
            for i, s in enumerate(self.sessions):
                try:
                    res = self.tasks[i].result()
                    result.append(res)
                except asyncio.exceptions.InvalidStateError:
                    pass
            if len(result) >= len(self.sessions):
                return result
            await asyncio.sleep(1)

    async def New2fa(self, code, logger):
        for i, s in enumerate(self.sessions):
            self.tasks.append(asyncio.create_task(s.New2fa(code, logger)))
        self.progress_bar.hint = "Установка 2FA 🔐"
        while True:
            result = await self._logging_task('new2fa')
            if result is True:
                break
            await asyncio.sleep(1)

    async def SetChannel(self, channel):
        for i, s in enumerate(self.sessions):
            self.tasks.append(asyncio.create_task(s.SetChannel(channel)))
        self.progress_bar.hint = TASK_NAMES['set_channel']
        while True:
            result = await self._logging_task('new2fa')
            if result is True:
                break
            await asyncio.sleep(1)

    async def CallPhone(self, wait, users, video):
        for i, chunk in enumerate(users):
            try:
                s = self.sessions[i]
            except IndexError:
                break
            self.tasks.append(asyncio.create_task(
                s.CallPhone(wait, chunk, video)))
            self.progress_bar.hint = TASK_NAMES['callphone']
        while True:
            result = await self._logging_task('callphone')
            await self._task_start()
            if result is True:
                break
            await asyncio.sleep(100000)

    async def WireTapping(self, wait, groups, keywords, admin):
        for i, chunk in enumerate(groups):
               try:
                    s = self.sessions[i]
                except IndexError:
                    break
                self.tasks.append(asyncio.create_task(
                    s.WireTapping(wait, chunk, keywords, admin)))
                self.progress_bar.hint = TASK_NAMES['wiretapping']
        while True:
            result = await self._logging_task('wiretapping')
            await self._task_start()
            if result is True:
                break
            await asyncio.sleep(100000)


    async def Reaction1(self, urls, reactions, wait):

        for i, chunk in enumerate(urls):
            try:
                s = self.sessions[i]
            except IndexError:
                break
            self.tasks.append(asyncio.create_task(
                s.Reaction1(chunk, reactions, wait)))
            self.progress_bar.hint = TASK_NAMES['liker']
        while True:
            result = await self._logging_task('reaction')
            await self._task_start()
            if result is True:
                break
            await asyncio.sleep(100000)

    async def Reaction2(self, urls, reactions, time):
        for i, chunk in enumerate(urls):
            try:
                s = self.sessions[i]
            except IndexError:
                break
            self.tasks.append(asyncio.create_task(
                s.Reaction2(chunk, reactions, time)))
            self.progress_bar.hint = TASK_NAMES['liker']
        while True:
            result = await self._logging_task('reaction')
            await self._task_start()
            if result is True:
                break
            await asyncio.sleep(100000)

    async def Reaction3(self, urls, reactions, wait):
        for i, chunk in enumerate(urls):
            try:
                s = self.sessions[i]
            except IndexError:
                break
            self.tasks.append(asyncio.create_task(
                s.Reaction3(chunk, reactions, wait)))
            self.progress_bar.hint = TASK_NAMES['liker']
        while True:
            result = await self._logging_task('reaction')
            await self._task_start()
            if result is True:
                break
            await asyncio.sleep(100000)

    async def SuggestPhoto(self, wait, users, suggest_photo, text=None, photo=None, video=None, message_id=None, channel_id=None):
        for i, chunk in enumerate(users):
            try:
                s = self.sessions[i]
            except IndexError:
                break
            self.tasks.append(asyncio.create_task(
                s.SuggestPhoto(wait, chunk, suggest_photo, text, photo, video, message_id, channel_id)))
            self.progress_bar.hint = TASK_NAMES['suggest_photo']
        while True:
            result = await self._logging_task('suggestphoto')
            await self._task_start()
            if result is True:
                break
            await asyncio.sleep(100000)


    async def AddContacts(self, wait, users):
        for i, chunk in enumerate(users):
            try:
                s = self.sessions[i]
            except IndexError:
                break
            self.tasks.append(asyncio.create_task(
                s.AddContacts(wait, chunk)))
            self.progress_bar.hint = TASK_NAMES['add_contact']
        while True:
            result = await self._logging_task('addcontacts')
            await self._task_start()
            if result is True:
                break
            await asyncio.sleep(100000)


    async def PrepareAnswerMachine(self, text, url_answer_machine, wait):
        for i, s in enumerate(self.sessions):
            self.tasks_machine.append(asyncio.create_task(s.PrepareAnswerMachine(
                text, url_answer_machine, wait)))

    async def _logging_task(self, task_type):
        text = [""]
        is_ended = []

        data = await self.data["state"].get_data()

        users_in_work = data["users_in_work"]
        errors = data["errors"]
        successful = 0
        unsuccessful = 0

        for i, s in enumerate(self.sessions):
            try:

                error = self.tasks[i].result()
                if error is None:
                    try:
                        error = error[1] + " ❌"
                    except TypeError:
                        raise asyncio.InvalidStateError
                elif isinstance(error, list) and len(error) == 2:
                    print(error)
                    error = error[0]
                    error = error[1] + " ❌"
                elif (error is True or (error[0] is True and len(error) == 1)) and all(is_ended) and len(is_ended) == len(self.tasks):
                    error = "Завершено ✅"
                elif not isinstance(error, bool) and error[0] is True and len(error) == 2:
                    print(error)
                    error = error[1] + "✅"
                elif isinstance(error, list) and len(error) == 2 and error[0] is False:
                    print(error)
                    error = error[1] + " ❌"
                    return False
                elif isinstance(error, tuple) and len(error) == 2 and error[0] is None:
                    try:
                        error = error[1] + " ❌"
                    except:
                        print(error)
                        error = str(errors[1]) + " ❌"
                elif error is True and len(is_ended) == len(self.tasks):
                    error = "Завершено ✅"
                elif error is True and  len(is_ended) != len(self.tasks):
                    error = ""
                else:
                    print(error)
                    error = str(error[1]) + " ❌"

                if s.client.is_connected():
                    s.client.disconnect()

                is_ended.append(True)

                if error == '' and len(is_ended) == len(self.tasks):
                    error = "Завершено ✅"

            except asyncio.InvalidStateError:
                is_ended.append(False)

            except asyncio.CancelledError as e:
                error = "Отменено ✅"
                if s.client.is_connected():
                    s.client.disconnect()
                is_ended.append(True)

            except IndexError:
                continue
            except Exception as e:
                error = "Завершено ✅"
                self.logger.error(f"Ошибка: {e}")
                print(e)

            try:
                if "Завершено" not in error and "Отменено" not in error and error:
                    try:
                        if s.me is not None:
                            if s.me.id not in errors:
                                self.count_errors += 1
                                errors.append(s.me.id)
                                await self.data["state"].update_data(errors=errors)
                    except:
                        self.count_errors += 1
                        await self.data["state"].update_data(errors=errors)
            except UnboundLocalError:
                error = ""
            except Exception as e:
                print(e)

            if task_type != "reset" and task_type != "new2fa":
                await self.data["state"].update_data(users_in_work=users_in_work)
            if task_type == "masslooking" or task_type == "masslookingchats":
                successful += sum([i for i in s.result.values()])
            elif task_type == "parsing":
                try:
                    successful = self.data["successful"]
                except:
                    successful = 0
            elif task_type == "reset":
                pass
            elif task_type == "number_check":
                successful = len([i for i in s.result.keys()])
                unsuccessful = sum(1 for i in s.result.values()
                                   if i is False or i is None)
            elif task_type == "new2fa":
                pass
            else:
                successful += sum(1 for i in s.result.values() if i is True)

                unsuccessful += sum(1 for i in s.result.values()
                                    if i is False or i is None)
                # if s.result.keys():
                #     self.remaining.append([item for item in s.users if item not in s.result.keys()])
            if task_type in ("mentions", "mailing", "wiretapping", "liker"):
                users_in_work = len(self.data["identifiers"])

            elif task_type == "masslooking":
                users_in_work = len(s.result.keys())
            elif task_type == "reset":
                pass
            elif task_type == "new2fa":
                pass
            else:
                users_in_work = successful+unsuccessful

        task_type == self.data["task_id"]
        try:
            error_s = error
        except UnboundLocalError:
            error = "❌ Функция была завершена некорректно"
        if task_type != "reset" and task_type != "new2fa":
            if task_type == "masslooking":
                   session_hint = f'⏱️ Задержка между просмотрами: {self.data["looking"][0]} - {self.data["looking"][1]} сек.\n\n'
                    session_hint += f'⏲ Задержка между простановкой реакции: {self.data["reaction_flood"][0]} - {self.data["reaction_flood"][1]} сек.\n\n'
            elif task_type == "masslookingchats":
                session_hint = ""
            elif task_type == "search_chats":
                session_hint = ""
            elif task_type == "tipping":
                session_hint = ""
            elif task_type == "number_check":
                view_delay = self.data["view_delay"].split("-")
                session_hint = f'⏱️ Задержка: {view_delay[0]} - {view_delay[1]} сек.\n\n'
            else:
                try:
                    session_hint = f'⏱️ Задержка: {self.data['wait'][0]} - {self.data['wait'][1]} сек.\n\n'
                except TypeError:
                    session_hint = f'⏱️ Задержка: {self.data['entry'][0]} - {self.data['entry'][1]} сек.\n\n'
                except KeyError:
                    pass

            session_hint += f'🔢 Используется аккаунтов: {len(self.sessions)}\n\n'
            session_hint += f'☠ Ошибок аккаунтов: {self.count_errors}\n\n'
        if task_type != "parsing" and task_type != "reset" and task_type != "search_chats" and task_type != "new2fa" and task_type != "number_check":
            session_hint += f'👨‍💼 Пользователей в списке: {len(self.data["identifiers"])}\n\n' if task_type not in (
                "mailing", "tipping", "neurocommenting", "commenting", "wiretapping", "liker", "mentions", "masslookingchats") else f'👨‍💼 Чатов в списке: {len(self.data["identifiers"])}\n\n'
        if task_type == "number_check":
            session_hint += f'📲 Номеров в списке: {len(self.data["user_ids"])}\n\n'
        if task_type == "spaming":
            session_status = f'Успешно: {successful}\nНеудачно: {unsuccessful}\n\nСтатус: {error if error and all(is_ended) else "⏳ Работает"} '
        elif task_type == "mentions":
            session_status = f'Успешно: {successful}\nНеудачно: {unsuccessful}\n\nСтатус: {error if error else "⏳ Работает"} '
        elif task_type == "parsing":
            for groupname, info in s.result.items():
                try:
                    successful += len(info.keys())
                    await self.data["state"].update_data(successful=successful)
                except:
                    pass
            session_status = f'✔ Спарсено: {successful}\n\nСтатус: {error if error else "⏳ Работает"} '
        elif task_type == "search_chats":
            for keyword, info in s.result.items():
                try:
                    successful += len(info)
                    await self.data["state"].update_data(successful=successful)
                except:
                    pass
            session_status = f'✔ Найдено: {successful}\n\nСтатус: {error if error else "⏳ Работает"} '
        elif task_type == "reports":
            session_status = f'Отправлено жалоб: {len(list(s.result.keys()))}\n\nСтатус: {error if error else "⏳ Работает"} '
        elif task_type == "inviting":
            session_status = f'Успешно: {successful}\nНеудачно: {unsuccessful}\n\nСтатус: {error if error else "⏳ Работает"} '
        elif task_type == "suggestphoto":
            session_status = f'Успешно: {successful}\nНеудачно: {unsuccessful}\n\nСтатус: {error if error else "⏳ Работает"} '
        elif task_type == "massteging":
            session_status = f'Успешно: {successful}\nНеудачно: {unsuccessful}\n\nСтатус: {error if error else "⏳ Работает"} '
        elif task_type == "addcontacts":
            session_status = f'Успешно: {successful}\nНеудачно: {unsuccessful}\n\nСтатус: {error if error else "⏳ Работает"} '
        elif task_type == "callphone":
            session_status = f'Успешно: {successful}\nНеудачно: {unsuccessful}\n\nСтатус: {error if error else "⏳ Работает"} '
        elif task_type == "tipping":
            session_status = f'Успешно: {successful}\nНеудачно: {unsuccessful}\n\nСтатус: {error if error else "⏳ Работает"} '
        elif task_type == "mailing":
            session_status = f'Успешно: {successful}\nНеудачно: {unsuccessful}\n\nСтатус: {error if error else "⏳ Работает"} '
        elif task_type == "masslooking":
            session_status = f'✔ Просмотрено: {successful}\n\nСтатус: {error if error else "⏳ Работает"} '
        elif task_type == "masslookingchats":
            session_status = f'✔ Просмотрено: {successful}\n\nСтатус: {error if error else "⏳ Работает"} '
        elif task_type == "reaction":
            session_status = f'Статус: {"Работает⏳" if not error else error}'
        elif task_type == "neurocommenting":
            session_status = f'Успешно: {successful}\nНеудачно: {unsuccessful}\n\nСтатус: {error if error else "⏳ Работает"} '
        elif task_type == "commenting":
            session_status = f'Успешно: {successful}\nНеудачно: {unsuccessful}\n\nСтатус: {error if error else "⏳ Работает"} '
        elif task_type == "reset":
            session_hint = ""
            session_status = f'Статус: {error if error else "⏳ Работает"} '
        elif task_type == "number_check":
            session_status = f'Успешно: {successful}\nНеудачно: {unsuccessful}\n\nСтатус: {error if error else "⏳ Работает"} '
        elif task_type == "new2fa":
            session_hint = ""
            session_status = f'Статус: {error if error else "⏳ Работает"} '
        elif task_type == "wiretapping":
            session_status = f'Статус: {error if error else "⏳ Работает"} '
        if task_type != "parsing" and task_type != "masslooking" and task_type != "masslookingchats" and task_type != "reset" and task_type != "search_chats" and "number_check":
            session_hint += f'⚒ Пользователей в работе: {users_in_work}\n\n' if task_type not in ("mailing", "tipping", "neurocommenting", "wiretapping", "liker", "mentions") else f'⚒ Чатов в работе: {users_in_work}\n\n'
        elif task_type == "number_check":
            session_hint += f'⚒ Номеров в работе: {users_in_work}\n\n'
        elif task_type == "masslooking" or task_type == "masslookingchats":
            session_hint += f"⚒ Пользователей в работе: {users_in_work}\n\n"

        text.append(f'{session_hint} {session_status}')

        if all(is_ended):
            if task_type == "spaming":

                self.logger.info("Всего отправлено: {}".format(successful))
                for i, s in enumerate(self.sessions):

                    self.logger.info(f'{s.session.replace(".session", "").replace(
                        f"{self.work_dir}/", "")} - {len([i for i in s.result.values() if i is True])} отправлено')
                dir = self.path
                result = []
                for sublist in self.remaining:

                    result.append("@" + sublist)
                with open(dir+"/remaining.txt", "w", encoding="utf-8") as file:
                    file.write("\n".join(result))

            elif task_type == "mentions":
                self.logger.info("Всего отмечено: {}".format(successful))
                for i, s in enumerate(self.sessions):
                    self.logger.info(f'{s.session.replace(".session", "").replace(
                        f"{self.work_dir}/", "")} - {len([i for i in s.result.values() if i is True])} отмечено')
            elif task_type == "massteging":
                self.logger.info("Всего тегнуто: {}".format(successful))
                for i, s in enumerate(self.sessions):
                    self.logger.info(f'{s.session.replace(".session", "").replace(
                        f"{self.work_dir}/", "")} - {len([i for i in s.result.values() if i is True])} тегнуто')
            elif task_type == "masslooking" or task_type == "masslookingchats":
                self.logger.info("Всего просмотрено: {}".format(successful))
                for i, s in enumerate(self.sessions):
                    self.logger.info(f'{s.session.replace(".session", "").replace(
                        f"{self.work_dir}/", "")} - {sum([i for i in s.result.values()])} просмотрено')
            elif task_type == "callphone":
                self.logger.info("Всего звонков: {}".format(successful))
                for i, s in enumerate(self.sessions):
                    self.logger.info(f'{s.session.replace(".session", "").replace(
                        f"{self.work_dir}/", "")} - {len([i for i in s.result.values() if i is True])} звонков')
            elif task_type == "suggestphoto":
                self.logger.info("Всего предложено: {}".format(successful))
                for i, s in enumerate(self.sessions):
                    self.logger.info(f'{s.session.replace(".session", "").replace(
                        f"{self.work_dir}/", "")} - {len([i for i in s.result.values() if i is True])} предложено')
            elif task_type == "inviting":
                self.logger.info("Всего приглашено: {}".format(successful))
                for i, s in enumerate(self.sessions):
                    self.logger.info(f'{s.session.replace(".session", "").replace(
                        f"{self.work_dir}/", "")} - {len([i for i in s.result.values() if i is True])} приглашено')
            elif task_type == "mailing":
                self.logger.info("Всего отправлено: {}".format(successful))
                for i, s in enumerate(self.sessions):
                    self.logger.info(f'{s.session.replace(".session", "").replace(
                        f"{self.work_dir}/", "")} - {len([i for i in s.result.values() if i is True])} отправлено')
            try:
                time = self.id.split("=")[1]

                if not os.path.exists(f"tasks/{self.data["customer"].id}/{self.task_id}"):
                    os.mkdir(
                        f"tasks/{self.data['customer'].id}/{self.task_id}/{time}")
                with open(f"tasks/{self.data['customer'].id}/{self.task_id}/{time}/progressbar.txt", "w", encoding="utf-8") as f:
                    f.write("\n".join(text))

            except:
                pass
            for task in data["accounts"]:
                if task.id == self.id:
                    data["accounts"].remove(task)
                    try:
                        del active_tasks[task.task_id]
                    except:
                        pass
                    await self.data["state"].update_data(accounts=data["accounts"])
                    for session in task.sessions:

                        for callback, event_ in session.client.list_event_handlers():
                            session.client.remove_event_handler(event_)
                            session.client.remove_event_handler(callback)
            await self.progress_bar.Success(progress.TASK_PROGRESS.format(info="\n".join(text)), reply_markup=kb.get_to_tasks_button_new(self.id, self.path))

            return True
        try:
            callback = data["callback"]
        except Exception as e:
            callback = None
        if task_type in ("parsing", "liker", "wiretapping"):
            await self.progress_bar.Process(progress.TASK_PROGRESS.format(info="\n".join(text)), reply_markup=kb.get_cancel_keyboard(self.id, self.path), callback=callback)
        elif task_type == "2fa":
            await self.progress_bar.Process(progress.TASK_PROGRESS.format(info="\n".join(text)), reply_markup=kb.wait_button, callback=callback)
        else:
            await self.progress_bar.Process(progress.TASK_PROGRESS.format(info="\n".join(text)), reply_markup=kb.get_cancel_keyboard(self.id, self.path), callback=callback)

    async def _task_start(self):
        await self.progress_bar.msg.edit_text(STARTED_TASK, reply_markup=kb.get_started_keyboard(self.id))

    async def StopAll(self):
        for session in self.sessions:
            if type(session.client) is not str:
                if session.client.is_connected():

                    try:
                        await session.client.disconnect()
                    except OperationalError:
                        pass

    async def CancelAll(self):
        for task in self.tasks:
            task.cancel()
        for task in self.tasks_machine:
            task.cancel()


class Session:
    def __init__(self, session: str, proxy: str = None, flood_wait: int = None, logger= None):
        self.proxy = proxy
        self.session = session
        self.result = {}
        self.users = []
        self.urls = []
        self.remaining = []

        self.logger = logger if logger is not None else logging.getLogger()
        self.spamblock = None
        self.me = None
        self.params = self.GetParams()
        self.open_api_client = OpenAI(
            api_key=OPEN_AI_TOKEN,
            base_url="http://localhost:1337/v1"
        )
        if self.params is None or (self.params.get("api_id") is None or self.params.get("api_hash") is None):
            self.client = "ParametersError"

        else:
            self.twofa = self.params.get("2fa")
            self.params.pop('2fa')
            try:
                from telethon.network.connection.tcpabridged import ConnectionTcpAbridged
                from telethon.network.connection.tcpobfuscated import ConnectionTcpObfuscated
                self.client = TelegramClient(
                    session,
                    proxy=ProxyFromUrl(proxy) if proxy is not None else None,
                    system_version="4.16.30-vxCUSTOM",

                )

            except Exception as e:
                self.client = "ClientError"
            else:
                if flood_wait is not None:
                    self.client.flood_sleep_threshold = flood_wait
                else:
                    self.client.flood_sleep_threshold = 0

    def GetParams(self):
        if not os.path.exists(self.session.replace(".session", ".json")):
            return

        with open(self.session.replace(".session", ".json"), mode='r', encoding='utf-8') as f:
            try:
                json_data = json.loads(f.read())
            except (json.decoder.JSONDecodeError, UnicodeDecodeError):
                return

            return {
                'api_id': self._get_value(json_data, 'api_id', 'app_id', 'apiId', 'appId'),
                'api_hash': self._get_value(json_data, 'api_hash', 'app_hash', 'apiHash', 'appHash'),
                'device_model': self._get_value(json_data, 'deviceModel', 'device'),
                'system_version': self._get_value(json_data, 'systemVersion', 'system_version', 'appVersion', 'app_version'),
                'app_version': self._get_value(json_data, 'appVersion', 'app_version'),
                'lang_code': self._get_value(json_data, 'lang_pack', 'langPack', 'lang_code', 'langCode'),
                'system_lang_code': self._get_value(json_data, 'system_lang_pack', 'systemLangPack', 'system_lang_code', 'systemLangCode'),
                '2fa': self._get_value(json_data, 'twoFA', '2fa', '2FA', 'password'),
            }

    def _get_value(self, file_json, *keys):
        for key in keys:
            if key in file_json:
                return file_json[key]
        return None

    def is_russian(self, text):
        try:
            return detect(text) == 'ru'
        except:
            return False

    def distribute_urls(self, urls, num_accounts, chats_per_account):

        distributed_urls = [[] for _ in range(num_accounts)]
        for i, url in enumerate(urls):
            distributed_urls[i % num_accounts].append(url)
        return [urls[:chats_per_account] for urls in distributed_urls]

    async def AllAdmin(self, sessions, chat):
        for session in sessions:
            try:
                user = await self.client.get_entity(session.me.username)
            except:
                self.logger.error(f"Не смог получить юзера {session}")
            try:
                await self.client.edit_admin(chat, user, invite_users=True)
            except Exception as e:
                print(e)

    async def send_messages_to_chats(self, client, text,
                                     file_id, chat_groups, delay_send, counts_message):
        async def send_message(chat_id):
            try:
                chat_id = await client.get_entity(chat_id)
            except Exception as e:
                self.logger.error(
                    f'Клиент {self.me.phone} не смог получить информацию о чате {e}')

            try:
                if file_id is None:
                    self.logger.info(
                        f'Клиент {self.me.phone} отправил сообщение в {chat_id.title}')
                    res = await client.send_message(chat_id, text)
                    await asyncio.sleep(0.1)
                    self.result[res.id] = True
                    return True
                else:
                    file = "test.jpg"
                    self.logger.info(
                        f'Клиент {self.me.phone} отправил сообщение с файлом в {chat_id.title}')
                    res = await client.send_message(chat_id, text)
                    self.result[res.id] = True
                    return True
            except Exception as e:
                self.logger.error(
                    f'Клиент {self.me.phone} не смог отправить сообщение в {chat_id.title}: {e}')

        for _ in range(counts_message):
            tasks = [send_message(chat_id) for chat_id in chat_groups]
            results = await asyncio.gather(*tasks)

            await asyncio.sleep(delay_send)

        return True

    async def Start(self):
        if not self.client.is_connected():
            status = await self.Check()
            if status is not True and "ограничен" not in status:
                return None, status

    async def Check(self, spam_block: bool = False):
        # Подключение к аккаунту и запуск бота
        await asyncio.sleep(0)
        if self.client == "ParametersError":
            self.logger.critical(
                f"[{self.session.split('/')[-1]}] Не смог прочитать JSON файл")
            return None, "Не смог прочитать JSON файл"
        elif self.client == "ClientError":
            self.logger.critical(
                f"[{self.session.split('/')[-1]}] Не смог прочитать SESSION файл(проверьте, точно ли он для Telethon)")
            return None, "Не смог прочитать SESSION файл(проверьте, точно ли он для Telethon"

        try:
            self.logger.info(
                f"[{self.session.split('/')[-1]}] Подключение к аккаунту...")
            try:
                await asyncio.wait_for(self.client.connect(), timeout=60.0)
            except asyncio.TimeoutError:
                self.logger.critical(
                    f"[{self.session.split('/')[-1]}] Таймаут подключения")
                return "Таймаут подключения"

            if not await self.client.is_user_authorized():
                await self.client.disconnect()
                self.logger.critical(
                    f"[{self.session.split('/')[-1]}] Аккаунт забанен")
                return "Аккаунт забанен"
            try:
                self.me = await self.client.get_me()
            except (errors.rpcerrorlist.AuthKeyDuplicatedError, errors.common.TypeNotFoundError):
                await self.client.disconnect()
                raise errors.rpcerrorlist.AuthKeyDuplicatedError("Error")
            except errors.rpcerrorlist.FloodWaitError:
                self.spamblock = True
                self.logger.warning(
                    f"[{self.session.split('/')[-1]}] Аккаунт имеет ограничения")
                return "Аккаунт имеет ограничения"
        except ConnectionError:
            self.logger.critical(
                f"[{self.session.split('/')[-1]}] Подключение неожиданно прервано")
            return "Подключение неожиданно прервано"
        except (TypeError) as e:
            self.logger.critical(
                f"[{self.session.split('/')[-1]}] Ошибка авторизации в прокси-сервере")
            return "Ошибка авторизации в прокси-сервере, запустите чекер еще раз или проверьте свой прокси"
        except OperationalError as e:
            self.logger.critical(
                f"[{self.session.split('/')[-1]}] Подключен из другого места")
            return "Подключен из другого места"
        except errors.rpcerrorlist.AuthKeyDuplicatedError:
            self.logger.critical(
                f"[{self.session.split('/')[-1]}] Подключен с другого IP адреса")
            return "Подключен с другого IP-адреса"
        except Exception as e:
            self.logger.critical(f"[{self.session.split('/')[-1]}] {e}")

        if spam_block:
            try:
                await self.client(functions.contacts.UnblockRequest('https://t.me/SpamBot'))
            except errors.rpcerrorlist.FloodWaitError:
                self.spamblock = True
                self.logger.warning(
                    f"[{self.session.split('/')[-1]}] Аккаунт имеет ограничения")
                return "Аккаунт имеет ограничения"
            except errors.rpcbaseerrors.FloodError:
                return "Аккаунт заморожен"

            async with self.client.conversation('https://t.me/SpamBot') as conv:
                await conv.send_message('/start')
                try:
                    msg = await conv.get_response()
                except Exception as e:

                    self.logger.warning(
                        f"[{self.session.split('/')[-1]}] Аккаунт используется другим")
                    return "Аккаунт используется другим"

            if 'UTC' in msg.text:
                self.spamblock = True
                self.logger.warning(
                    f"[{self.session.split('/')[-1]}] Аккаунт имеет ограничения")
                return "Аккаунт имеет ограничения"
            if 'limited' in msg.text or "antispam" in msg.text or 'abnormal' in msg.text or "Ваш аккаунт ограничен" in msg.text or "phone numbers" in msg.text:
                self.spamblock = True
                self.logger.warning(
                    f"[{self.session.split('/')[-1]}] Аккаунт имеет ограничения")
                return "Аккаунт имеет ограничения"
        try:
            res = await self.client(functions.account.UpdateStatusRequest(
                offline=False
            ))
        except:
            pass
        try:
            await self.client(functions.contacts.UnblockRequest('https://t.me/SpamBot'))
        except errors.rpcbaseerrors.FloodError:
            self.logger.warning(
                f"[{self.session.split('/')[-1]}] Аккаунт заморожен")

            return "Аккаунт заморожен"
        self.logger.warning(
            f"[{self.session.split('/')[-1]}] Аккаунт подключен!")
        # await self.client.disconnect()
        return True

    async def GetConfirmationCode(self):
        try:
            async for message in self.client.iter_messages(777000, limit=1):
                match = re.search(r'(\d+)', message.message)
                if match:
                    return True, {
                        'code': match.group(1),
                        '2fa': self.twofa if self.twofa is not None else "-"
                    }  # Extracting code from the message
            return None, "Сообщение с кодом не найдено"
        except errors.rpcerrorlist.FloodWaitError:
            return None, f"Флуд-задержка превысила {self.client.flood_sleep_threshold} сек."
        except errors.rpcerrorlist.AuthKeyDuplicatedError:
            return None, "Авторизован с другого IP-адреса"
        except (errors.rpcerrorlist.PhoneNumberInvalidError, errors.rpcerrorlist.PhoneNumberBannedError,
                errors.rpcerrorlist.PhonePasswordFloodError, errors.rpcerrorlist.PhoneCodeInvalidError) as e:
            return None, str(e)
        except ConnectionError:
            return None, "Подключение было прервано"
        except Exception as e:
            return None, str(e)

    async def SendMessage(self, user, text=None, video=None, photo=None, via_bot= None, sticker = None):
        try:
            while True:
                try:
                    if photo is not None:
                        await self.client.send_file(user, file=photo, caption=text, parse_mode='html')
                        self.logger.info(
                            f"Отправка сообщения пользователю {user}")
                    elif video is not None:
                        await self.client.send_file(user, file=video, caption=text, parse_mode='html', supports_streaming=True)
                        self.logger.info(
                            f"Отправка сообщения пользователю {user}")
                    elif via_bot is not None:
                        query = await self.client.inline_query("@PostBot", via_bot)
                        result = await query[0].click(user)
                        self.logger.info(
                            f"Отправка сообщения пользователю {user}")
                    elif sticker is not None:
                        await self.client.send_file(user, file=sticker)
                        self.logger.info(
                            f"Отправка сообщения пользователю {user}")
                    else:

                        await self.client.send_message(user, text, parse_mode='html')
                        self.logger.info(
                            f"Отправка сообщения пользователю {user}")
                    break
                except errors.rpcerrorlist.SlowModeWaitError as e:
                    await asyncio.sleep(e.seconds)
                    continue
        except (
            errors.rpcerrorlist.UserIsBlockedError, errors.rpcerrorlist.ChatAdminRequiredError,
            errors.rpcerrorlist.ChannelPrivateError, errors.rpcerrorlist.ChatWriteForbiddenError,
            errors.rpcerrorlist.InputUserDeactivatedError, errors.rpcerrorlist.TimeoutError,
            ValueError, errors.rpcerrorlist.UsernameInvalidError,
            errors.rpcerrorlist.ChatIdInvalidError, errors.rpcerrorlist.ChatRestrictedError, errors.rpcerrorlist.EntityBoundsInvalidError,
            errors.rpcerrorlist.EntityMentionUserInvalidError, errors.rpcerrorlist.PeerIdInvalidError, errors.rpcerrorlist.EntityBoundsInvalidError,
            OperationalError
        ) as e:
            self.logger.error(e.__class__.__name__)
            self.logger.error(f"error: {str(e)}")
        except errors.rpcerrorlist.PeerFloodError:
            self.logger.error(
                f"Аккаунт {self.me.phone} ограничен при попытке отправить сообщение пользователю {user}")

        except errors.rpcerrorlist.AuthKeyDuplicatedError:
            return None, "Авторизован с другого IP-адреса"
        except errors.rpcerrorlist.FloodWaitError as e:
            return None, f"Флуд-задержка превысила {self.client.flood_sleep_threshold} сек."
        except (ConnectionError):
            return None, "Подключение было оборвано"
        except Exception as e:
            self.logger.error(f"error: {str(e)}")
        else:
            return True

    async def SendMessageMentions(self, user, text=None, video=None, photo=None, usernames=None):
        try:
            while True:
                try:
                    if photo is not None:
                        await self.client.send_file(user, file=photo, caption=text, parse_mode='html')
                        self.logger.info(
                            f"Отметил пользователей {','.join(str(username) for username in usernames)} в {user}")
                        for username in usernames:
                            self.result[username] = True
                    elif video is not None:
                        await self.client.send_file(user, file=video, caption=text, parse_mode='html', supports_streaming=True)
                    else:

                        result = await self.client.send_message(user, text, parse_mode='html')
                        self.logger.info(f"Отметил пользователей {','.join(
                            str(username) for username in usernames)} в {user}")
                        for username in usernames:
                            self.result[username] = True
                    break
                except errors.rpcerrorlist.SlowModeWaitError as e:
                    await asyncio.sleep(e.seconds)
                    continue
        except (
            errors.rpcerrorlist.UserIsBlockedError, errors.rpcerrorlist.ChatAdminRequiredError,
            errors.rpcerrorlist.ChannelPrivateError, errors.rpcerrorlist.ChatWriteForbiddenError,
            errors.rpcerrorlist.InputUserDeactivatedError, errors.rpcerrorlist.TimeoutError,
            ValueError, errors.rpcerrorlist.UsernameInvalidError,
            errors.rpcerrorlist.ChatIdInvalidError, errors.rpcerrorlist.ChatRestrictedError, errors.rpcerrorlist.EntityBoundsInvalidError,
            errors.rpcerrorlist.EntityMentionUserInvalidError, errors.rpcerrorlist.PeerIdInvalidError, errors.rpcerrorlist.EntityBoundsInvalidError,
            OperationalError
        ) as e:
            self.logger.error(f"error: {str(e)}")
            return None, str(e)
        except errors.rpcerrorlist.PeerFloodError:
            self.logger.error(f"Аккаунт {self.me.phone} ограничен")
            return None

        except errors.rpcerrorlist.AuthKeyDuplicatedError:
            return None, "Авторизован с другого IP-адреса"
        except errors.rpcerrorlist.FloodWaitError:
            return None, f"Флуд-задержка превысила {self.client.flood_sleep_threshold} сек."
        except (ConnectionError):
            return None, "Подключение было оборвано"
        except Exception as e:
            self.logger.error(f"error: {str(e)}")
            return None
        else:
            return True

    async def ChangeInfo(self, firstname, lastname, description, username, photo, story):
        if not self.client.is_connected():
            status = await self.Check()
            if status is not True and "ограничен" not in status:
                return None, status
        try:
            photos =await self.client.get_profile_photos("me")

            input_photos = [InputPhoto(id=photo.id, access_hash=photo.access_hash,
                                       file_reference=photo.file_reference) for photo in photos]
            if photo is not None:
                file = await self.client.upload_file(photo)
            else:
                file = None
            if firstname is None:
                firstname = self.me.first_name
            if lastname is None:
                lastname = self.me.last_name
            if description is not None:
                await self.client(UpdateProfileRequest(
                    first_name=firstname,
                    last_name=lastname,
                    about=description,
                ))
            elif firstname is not None or lastname is not None:
                await self.client(UpdateProfileRequest(
                    first_name=firstname,
                    last_name=lastname
                ))
            if username is not None:
                try:
                    await self.client(UpdateUsernameRequest(
                        username=username.replace("@", "")
                    ))
                except errors.rpcerrorlist.UsernameOccupiedError:
                    self.logger.error(
                        f"Аккаунт {self.me.phone} при попытке поставить username: {username} занят!")
                    return "Имя пользователя уже занято"

            if story is not None:
                res = await self.client(functions.stories.SendStoryRequest(
                    peer=self.me,
                    media=types.InputMediaUploadedPhoto(
                            file=await self.client.upload_file(story),

                        ),
                    privacy_rules=[types.InputPrivacyValueAllowAll()],
                    pinned=True,
                    noforwards=True,
                    fwd_modified=True
                ))

            await self.client(DeletePhotosRequest(input_photos))
            if file is not None:
                await self.client(UploadProfilePhotoRequest(
                    file=file

                ))

            self.logger.info(
                f"Изменение информации профиля прошло успешно для аккаунта {username if username is not None else "@"+self.me.username if self.me.username is not None else self.me.phone}")
            await self.client.disconnect()

            return True
        except errors.AboutTooLongError:
            self.logger.error(
                f"Аккаунт {self.me.phone} при попытке поставить описание: описание слишком длинное! (максимум 70 символов)")
            return "Описание слишком длинное! (максимум 70 символов)"
        except errors.UsernameNotModifiedError:
            self.logger.error(
                f"Аккаунт {self.me.phone} при попытке поставить username: username не изменился!")
            return "Имя пользователя не изменилось!"
        except Exception as e:
            self.logger.error(f"error: {str(e)}")
            return str(e)

    async def RepostMessage(self, user, message_id=None, channel_id=None):
        try:
            await self.client.forward_messages(entity=user, from_peer=channel_id, messages=[message_id])
            self.logger.info(f"Пересылка сообщения пользователю {user}")
        except (
            errors.rpcerrorlist.UserIsBlockedError, errors.rpcerrorlist.ChatAdminRequiredError,
            errors.rpcerrorlist.ChatWriteForbiddenError, errors.rpcerrorlist.InputUserDeactivatedError,
            errors.rpcerrorlist.TimeoutError, ValueError,
            errors.rpcerrorlist.UsernameInvalidError
        ) as e:
            self.logger.error(f"error: {str(e)}")

        except (errors.rpcerrorlist.ChannelInvalidError, errors.rpcerrorlist.ChannelPrivateError, errors.rpcerrorlist.MessageIdInvalidError):
            return None, "Сообщение для пересылки не найдено"

        except errors.rpcerrorlist.FloodWaitError:
            return None, f"Флуд-задержка превысила {self.client.flood_sleep_threshold} сек."
        except errors.rpcerrorlist.PeerFloodError:
            self.logger.error(
                f"Аккаунт {self.me.phone} ограничен при попытке отправить сообщение пользователю {user}")
        except ConnectionError:
            return None, "Подключение было оборвано"

        except Exception as e:
            self.logger.error(f"error: {str(e)}")

        else:
            return True

    async def Spaming(self, users, wait, text=None, photo=None, video=None, message_id=None, channel_id=None, count=0, remaining=None, via_bot=None, sticker= None):
        self.users = users

        count_true = 0
        if type(text) == list:
            text = random.choice(text)
        if not self.client.is_connected():
            status = await self.Check()
            if status is not True and "ограничен" not in status:
                return None, status

        for i, user in enumerate(users):
            await asyncio.sleep(0)
            if channel_id or message_id:
                result = await self.RepostMessage(user, message_id, channel_id)
            elif sticker:
                sticker_sets = await self.client(GetAllStickersRequest(0))
                sticker_set = sticker_sets.sets[0]

                stickers = await self.client(GetStickerSetRequest(
                    stickerset=InputStickerSetID(
                        id=sticker_set.id, access_hash=sticker_set.access_hash
                    ),
                    hash=0
                ))
                sticker_ = random.choice(stickers.documents)
                result = await self.SendMessage(user, text, video, photo, via_bot, sticker_)
            else:
                if video is not None:
                    video = await self.client.upload_file(video)
                elif photo is not None:
                    photo = await self.client.upload_file(photo)
                result = await self.SendMessage(user, text, video, photo, via_bot)
            self.result[user] = result
            if result == True:
                count_true += 1
            if type(result) is tuple:
                return result
            if i < len(users)-1:
                await asyncio.sleep(rd(*wait))
        if remaining == []:
            return True
        while True:
            for i, user in enumerate(remaining):
                if count_true == count:
                    return True
                else:
                    if channel_id or message_id:
                        result = await self.RepostMessage(user, message_id, channel_id)
                    else:
                        result = await self.SendMessage(user, text, video, photo, via_bot)
                    self.result[user] = result
                    if result == True:
                        count_true += 1
                        remaining.remove(user)
                    if type(result) is tuple:
                        return result
                    if count_true == count:
                        return True
                    if i < len(remaining)-1:
                        await asyncio.sleep(rd(*wait))
            return True, f"Пользователи в базе закончились"

    async def Reports(self, wait, chats, report_text, report_reason):
        self.users = chats

        if not self.client.is_connected():
            status = await self.Check()
            if status is not True and "ограничен" not in status:
                return None, status

            for i, chat in enumerate(chats):
                await asyncio.sleep(0)
                try:
                    await self.client(functions.messages.ReportRequest(
                        peer=chat,
                        id=[i],
                        reason=report_reasons[report_reason],
                        message=report_text
                    ))
                    self.logger.info(
                        f"Отправка жалобы на {chat} (причина: {report_reasons[report_reason]} сообщение: {report_text})")
                except (errors.ChannelInvalidError, ValueError):
                    self.logger.info(f"Канал {chat} не существует")
                except errors.rpcerrorlist.FloodWaitError:
                    return None, f"Флуд-задержка превысила {self.client.flood_sleep_threshold} сек."
                except errors.rpcerrorlist.AuthKeyDuplicatedError:
                    return None, "Авторизован с другого IP-адреса"
                except ConnectionError:
                    return None, "Подключение было оборвано"
                except UsernameInvalidError:
                    self.logger.info(f"Пользователь {chat} не существует")
                except PeerIdInvalidError:
                    self.logger.info(f"Пользователь {chat} не существует")
                except Exception as e:
                    self.logger.error(f"error: {str(e)}")
                self.result[chat] = True

        return True

    async def Tipping(self, urls):
        if not self.client.is_connected():
            status = await self.Check()
            if status is not True and "ограничен" not in status:
                return None, status

        self.urls = urls
        for url in urls:
            try:
                if "joinchat" in url or "+" in url:
                    channel_id = url.split('/')[-1].replace('+', '')
                    await self.client(ImportChatInviteRequest(channel_id))
                else:
                    channel = await self.client.get_entity(url)
                    await self.client(JoinChannelRequest(channel))


                self.logger.info(
                    f'Клиент {self.me.phone} присоединился к {url}')
                self.result[url] = True

            except UserAlreadyParticipantError:
                self.logger.info(
                    f'Клиент {self.me.phone} уже состоит в {url}')
                self.result[url] = True
            except Exception as e:
                self.logger.error(
                    f'Клиент {self.me.phone} не смог присоединиться к {url}: {e}')
                self.result[url] = False

        return True

    async def FetchMailing(self, streams, delay_join, delay_send, counts_message, count_mailings, delay_mailing, file_id, text, urls):
        if not self.client.is_connected():
            status = await self.Check()
            if status is not True and "ограничен" not in status:
                return None, status
        self.urls = urls
        for url in urls:
            try:
                if "joinchat" in url or "+" in url:
                    channel_id = url.split('/')[-1].replace('+', '')
                    await self.client(ImportChatInviteRequest(channel_id))
                else:
                    channel = await self.client.get_entity(url)
                    await self.client(JoinChannelRequest(channel))
                self.logger.info(
                    f'Клиент {self.me.phone} присоединился к {url.replace("\r", "")}')
            except UserAlreadyParticipantError:
                pass
            except Exception as e:
                self.logger.error(
                    f'Клиент {self.me.phone} не смог присоединиться к {url.replace("\r", "")}: {e}')
            await asyncio.sleep(delay_join)
        for _ in range(count_mailings):

            result = await self.send_messages_to_chats(self.client, text, file_id, urls, delay_send, counts_message)

            await asyncio.sleep(delay_mailing)

        return result

    async def AnswerMachine(self, url):
        if not self.client.is_connected():
            status = await self.Check()
            if status is not True and "ограничен" not in status:
                return None, status
        try:
            if "joinchat" in url or "+" in url:
                channel_id = url.split('/')[-1].replace('+', '')

                await self.client(ImportChatInviteRequest(channel_id))
                channel = await self.client.get_entity(url)
                chat_id = channel.id
            else:
                channel = await self.client.get_entity(url)
                await self.client(JoinChannelRequest(channel))
                chat_id = channel.id
            self.logger.info(f'Клиент {self.me.phone} присоединился к {url}')
        except UserAlreadyParticipantError:
            channel = await self.client.get_entity(url)
            chat_id = channel.id
        except Exception as e:
            self.logger.error(
                f'Клиент {self.me.phone} не смог присоединиться к {url}: {e}')
        self.client.add_event_handler(lambda event: message_handler(
            event, chat_id), events.NewMessage())

    async def PrepareAnswerMachine(self, text, url_answer_machine, wait):
        if not self.client.is_connected():
            status = await self.Check()
            if status is not True and "ограничен" not in status:
                return None, status
        url = url_answer_machine
        try:
            if "joinchat" in url or "+" in url:
                channel_id = url.split('/')[-1].replace('+', '')

                await self.client(ImportChatInviteRequest(channel_id))
                channel = await self.client.get_entity(url)
                chat_id = channel.id
            else:
                channel = await self.client.get_entity(url)
                await self.client(JoinChannelRequest(channel))
                chat_id = channel.id
            self.logger.info(f'Клиент {self.me.phone} присоединился к {url}')
        except UserAlreadyParticipantError:
            channel = await self.client.get_entity(url)
            chat_id = channel.id
        except Exception as e:
            self.logger.error(
                f'Клиент {self.me.phone} не смог присоединиться к {url}: {e}')
        self.client.add_event_handler(lambda event: prepare_answer(
            event, text, url_answer_machine, wait), events.NewMessage(func=lambda e: e.is_private))

    async def Neurocommenting(self, channels, promt, url_answer_machine, type, entry, api):
        entry = list(map(int, entry.split("-")))
        if not self.client.is_connected():
            status = await self.Check()
            if status is not True and "ограничен" not in status:
                return None, status

        try:
            if "joinchat" in url_answer_machine or "+" in url_answer_machine:
                channel_id = url_answer_machine.split('/')[-1].replace('+', '')

                await self.client(ImportChatInviteRequest(channel_id))
                channel = await self.client.get_entity(url_answer_machine)
                chat_id = channel.id
            else:
                channel = await self.client.get_entity(url_answer_machine)
                await self.client(JoinChannelRequest(channel))
                chat_id = channel.id
            self.logger.info(
                f'Клиент {self.me.phone} присоединился к {url_answer_machine}')
            # await asyncio.sleep(flood)
        except UserAlreadyParticipantError:
            channel = await self.client.get_entity(url_answer_machine)
            chat_id = channel.id
        except Exception as e:
            self.logger.error(
                f'Клиент {self.me.phone} не смог присоединиться к {url_answer_machine}: {e}')
        try:
            if type == "latest":
                last_message_ids = {name: 0 for name in channels}
                for name in channels:
                    try:
                        channel_entity = await self.client.get_entity(name)
                    except ValueError as e:
                        await self.client.send_message(chat_id, f"Ошибка при получении информации о канале '{name}': {e}")
                        continue
                    messages = await self.client.get_messages(channel_entity, limit=5)
                    if messages:
                        for post in messages:
                            if post.id != last_message_ids[name]:
                                last_message_ids[name] = post.id
                                prompt = promt + post.raw_text if post.raw_text else promt
                                text = await promt_generation(api, prompt)

                                output = text

                                try:
                                    result = await self.client(functions.channels.GetFullChannelRequest(
                                        channel=channel_entity
                                    ))
                                    try:
                                        if result.full_chat.linked_chat_id is None:
                                            self.logger.info(
                                                f"Аккаунт {self.me.phone} не может оставить комментарий в {name} комментарии запрещены")
                                            self.result[post.id] = True
                                            break
                                        else:
                                            res = await self.client(JoinChannelRequest(result.full_chat.linked_chat_id))
                                    except Exception as e:
                                        self.logger.error(e)
                                    if "Произошла ошибка при генерации текста" not in output:
                                        res = await self.client.send_message(entity=channel_entity, message=output, comment_to=post.id)
                                        await self.client.send_message(chat_id, f"Отправлен комментарий: {output} в {name}")
                                        self.logger.info(
                                            f"Аккаунт {self.me.phone} отправил комментарий: {output} в {name}")
                                    self.result[post.id] = True

                                    await asyncio.sleep(random.randint(entry[0], entry[1]))
                                    break
                                except MsgIdInvalidError:
                                    continue
                                except ChannelPrivateError:
                                    self.logger.error(
                                        f"Аккаунт {self.me.phone} не может оставить комментарий в {name} Аккаунт был заблокирован в чате или у аккаунта нет прав, чтобы писать в чате")
                                except Exception as e:
                                    self.logger.error(e)
                                    await self.client.send_message(chat_id, f"Ошибка: {e}")
                                    self.logger.error(
                                        'Ошибка, проверьте сообщения в группе')
                return True
            else:
                for name in channels:
                    try:
                        channel_entity = await self.client.get_entity(name)
                    except ValueError as e:
                        await self.client.send_message(chat_id, f"Ошибка при получении информации о канале '{name}': {e}")
                        continue
                    messages = await self.client.get_messages(channel_entity, limit=1)
                    if messages:
                        for post in messages:
                            prompt = promt + post.raw_text
                            text = await promt_generation(api, prompt)
                            end = False
                            output = text
                            id = post.id
                            while True:
                                async for message in self.client.iter_messages(channel_entity, 1):
                                    if message.id > id:
                                        prompt = promt + post.raw_text
                                        text = await promt_generation(api, prompt)
                                        output = text
                                        result = await self.client(functions.channels.GetFullChannelRequest(
                                            channel=channel_entity
                                        ))
                                        res = await self.client(JoinChannelRequest(result.full_chat.linked_chat_id))
                                        id = await post_handler(self, chat_id, output, channel_entity)
                                        self.result[id] = True
                                        id = message.id
                                        # end = True

                                if end:
                                    break
                                await asyncio.sleep(10)
                    return True
        except errors.rpcerrorlist.FloodWaitError:
            return None, f"Флуд-задержка превысила {self.client.flood_sleep_threshold} сек."
        except Exception as e:
            self.logger.error(e)


    async def Commenting(self, channels, wait, type, text=None, photo=None, video=None, message_id=None, channel_id=None, via_bot=None, sticker= None):
        if not self.client.is_connected():
            status = await self.Check()
            if status is not True and "ограничен" not in status:
                return None, status

        if type == "latest":
            last_message_ids = {name: 0 for name in channels}
            for name in channels:
                channel_entity = await self.client.get_entity(name)
                messages = await self.client.get_messages(channel_entity, limit=5)
                if messages:
                    for post in messages:
                        if post.id != last_message_ids[name]:
                            last_message_ids[name] = post.id
                            output = text

                            try:
                                result = await self.client(functions.channels.GetFullChannelRequest(
                                    channel=channel_entity
                                ))
                                try:
                                    res = await self.client(JoinChannelRequest(result.full_chat.linked_chat_id))
                                except Exception as e:
                                    self.logger.error(e)
                                await self.client.send_message(entity=channel_entity, message=output, comment_to=post.id)
                                self.logger.info(
                                    f"Аккаунт {self.me.phone} отправил комментарий: {output} в {name}")
                                self.result[post.id] = True
                                break
                            except MsgIdInvalidError:
                                continue
                            except Exception as e:
                                self.logger.error(e)
                                self.logger.error(
                                    'Ошибка, проверьте сообщения в группе')
            return True
        else:
            for name in channels:
                chat_id = 0
                try:
                    channel_entity = await self.client.get_entity(name)
                except ValueError as e:
                    await self.client.send_message(chat_id, f"Ошибка при получении информации о канале '{name}': {e}")
                    continue
                messages = await self.client.get_messages(channel_entity, limit=1)
                # result = await self.client(functions.channels.GetFullChannelRequest(
                #                         channel=channel_entity
                #                     ))
                # res = await self.client(JoinChannelRequest(result.full_chat.linked_chat_id))
                if messages:
                    for post in messages:
                        end = False
                        output = text
                        id = post.id
                        while True:
                            async for message in self.client.iter_messages(channel_entity, 1):
                                if message.id > id:

                                    await post_handler(self, chat_id, output, channel_entity)
                                    self.result[id] = True
                                    id = message.id
                                    # end = True
                            if end:

                                break
                            await asyncio.sleep(10)
            return True

    async def CommentingEdit(self, channels, wait, entry, type_comment, first_comment, text=None, photo=None, video=None, message_id=None, channel_id=None, via_bot=None, sticker=None):
        if not self.client.is_connected():
            status = await self.Check()
            if status is not True and "ограничен" not in status:
                return None, status

        if type_comment == "latest":
            last_message_ids = {name: 0 for name in channels}
            for name in channels:
                channel_entity = await self.client.get_entity(name)
                messages = await self.client.get_messages(channel_entity, limit=5)
                if messages:
                    for post in messages:
                        if post.id != last_message_ids[name]:
                            last_message_ids[name] = post.id

                            try:
                                result = await self.client(functions.channels.GetFullChannelRequest(
                                    channel=channel_entity
                                ))
                                try:
                                    res = await self.client(JoinChannelRequest(result.full_chat.linked_chat_id))
                                except Exception as e:
                                    self.logger.error(e)
                                if type(first_comment) == list:
                                    first_comment = random.choice(
                                        first_comment)
                                if type(text) == list:
                                    text = random.choice(text)
                                message = await self.client.send_message(entity=channel_entity, message=first_comment, comment_to=post.id)

                                asyncio.create_task(edit_message_time(
                                    self, entry, message.id, result.full_chat.linked_chat_id, name, text))
                                await asyncio.sleep(*wait)
                                self.logger.info(
                                    f"Аккаунт {self.me.phone} отправил комментарий: {first_comment} в {name}")
                                self.result[post.id] = True
                                break
                            except MsgIdInvalidError:
                                continue
                            except Exception as e:
                                self.logger.error(e)
                                self.logger.error(
                                    'Ошибка, проверьте сообщения в группе')
            return True
        else:
            for name in channels:
                chat_id = 0
                try:
                    channel_entity = await self.client.get_entity(name)
                except ValueError as e:
                    await self.client.send_message(chat_id, f"Ошибка при получении информации о канале '{name}': {e}")
                    continue
                messages = await self.client.get_messages(channel_entity, limit=1)
                if messages:
                    for post in messages:
                        end = False
                        id = post.id
                        while True:
                            async for message in self.client.iter_messages(channel_entity, 1):
                                if message.id > id:
                                    result = await self.client(functions.channels.GetFullChannelRequest(
                                        channel=channel_entity
                                    ))
                                    res = await self.client(JoinChannelRequest(result.full_chat.linked_chat_id))
                                    post_id = await post_handler(self, chat_id, first_comment, channel_entity)
                                    asyncio.create_task(edit_message_time(
                                        self, entry, post_id, result.full_chat.linked_chat_id, name, text))
                                    await asyncio.sleep(*wait)
                                    self.result[message.id] = True
                                    id = message.id
                                    # end = True
                            if end:
                                break
                            await asyncio.sleep(10)
            return True

    async def MassLooking(self, reaction, looking, reaction_flood, stories_account, stories_user, identifiers, reaction_count):
        self.users = identifiers
        if not self.client.is_connected():
            status = await self.Check()
            if status is not True and "ограничен" not in status:
                return None, status
        count_of_reaction = 0
        flag = False
        total_view = 0
        for identifier in identifiers:
            try:

                if total_view >= stories_account:
                    self.result[identifier] = True
                    return True

                stories = await self.client(GetPeerStoriesRequest(
                    identifier,
                ))
                if len(stories.stories.stories) == 0:
                    self.logger.info(
                        f"Клиент {self.me.phone} не имеет историй {identifier}")
                    self.result[identifier] = 0
                    await asyncio.sleep(rd(*looking))
                    continue
                user = stories.users[0]
                if stories_user >= len(stories.stories.stories) or stories_user == 0:
                    stories = await self.client(ReadStoriesRequest(
                        identifier,
                        max_id=user.stories_max_id
                    ))
                    # if not flag:
                    #     self.result[identifier] += 1
                    # await self.logger.info(f"Клиент {self.me.phone} посмотрел историю {identifier} {story.id}")
                    # self.result[identifier] = len(stories)
                else:
                    stories_to_view = stories.stories.stories[:stories_user]
                    for story in stories_to_view:
                        stories = await self.client(ReadStoriesRequest(
                            user,
                            max_id=story.id
                        ))
                        # await self.logger.info(f"Клиент {self.me.phone} посмотрел историю {identifier} {story.id}")
                        # if not flag:
                        #     self.result[identifier] += 1
                    # self.result[identifier] = len(stories_to_view)

                await asyncio.sleep(rd(*looking))
                if stories == []:
                    self.logger.info(
                        f"Клиент {self.me.phone} уже просмотрел истории {identifier}")
                    # total_view += 1

                self.result[identifier] = 0
                if count_of_reaction < reaction_count:
                    for story_id in stories:
                        try:
                            await self.client(SendReactionRequest(
                                user,
                                story_id,
                                reaction=ReactionEmoji(
                                    emoticon=reaction,
                                )
                            ))
                            self.logger.info(
                                f"Клиент {self.me.phone} поставил реакцию истории {identifier} с идентификатором {story_id}")
                            total_view += 1
                            count_of_reaction += 1
                            self.result[identifier] += 1
                            await asyncio.sleep(rd(*reaction_flood))
                        except Exception as e:
                            self.logger.error(
                                f'Клиент {self.me.phone} не смог поставить реакцию истории {identifier}: {e}')
                else:

                    if not flag:
                        self.logger.info(
                            f"Клиент {self.me.phone} достигнул лимита реакций")
                        flag = True

            except Exception as e:
                self.logger.error(
                    f'Клиент {self.me.phone} не смог просмотреть истории {identifier}: {e}')
                await asyncio.sleep(rd(*looking))

        return True

    async def NumberCheck(self, wait, numbers):
        self.users = numbers
        if not self.client.is_connected():
            status = await self.Check()
            if status is not True and "ограничен" not in status:
                return None, status
        for number in numbers:
            try:
                contact = types.InputPhoneContact(
                    client_id=0, phone=number, first_name="", last_name=""
                )
                # Attempt to add the contact from the address book
                contacts = await self.client(functions.contacts.ImportContactsRequest([contact]))

                users = contacts.to_dict().get("users", [])
                number_of_matches = len(users)

                if number_of_matches == 0:
                    self.logger.info(
                        f"Клиент {self.me.phone} проверил номер {number} - НЕ ВАЛИД")
                    self.result[number] = False
                elif number_of_matches == 1:
                    # Attempt to remove the contact from the address book.
                    # The response from DeleteContactsRequest contains more information than from ImportContactsRequest
                    updates_response: types.Updates = await self.client(
                        functions.contacts.DeleteContactsRequest(
                            id=[users[0].get("id")])
                    )
                    user = updates_response.users[0]

                    if hasattr(user.status, 'was_online'):
                        status = user.status.was_online.strftime(
                            "%d.%m.%Y %H:%M:%S")
                    elif hasattr(user.status, 'expires') and not hasattr(user.status, 'was_online'):
                        status = "Онлайн"
                    else:
                        status = {
                            UserStatusEmpty: "Давно",
                            UserStatusLastWeek: "Был на прошлой неделе",
                            UserStatusLastMonth: "Был в прошлом месяце",
                            UserStatusRecently: "Был недавно"
                        }.get(type(user.status), "Был давно")
                    self.result[number] = {
                        "id": user.id,
                        "username": user.username,
                        "first_name": user.first_name,
                        "last_name": user.last_name,
                        "premium": user.premium,
                        "user_was_online": status,
                        "phone": number,
                    }
                    # getting more information about the user

                    self.logger.info(
                        f"Клиент {self.me.phone} проверил номер {number} - ВАЛИД")
                await asyncio.sleep(rd(*wait))
            except Exception as e:
                self.logger.error(
                    f'Клиент {self.me.phone} не смог добавить номер {number}: {e}')
                await asyncio.sleep(rd(*wait))
        return True

    async def MassLookingChats(self, reactions, chats, safe_mode, looking=[3, 5], reaction_flood=[3,5], wait=[3,5]):
        self.users = chats
        if not self.client.is_connected():
            status = await self.Check()
            if status is not True and "ограничен" not in status:
                return None, status

        async def masslook_chat(chat, looking, reaction_flood, reactions):
               try:
                    full = await self.client(functions.channels.GetFullChannelRequest(chat))
                except (errors.rpcerrorlist.ChannelPrivateError,
                        errors.rpcerrorlist.TimeoutError,
                        errors.rpcerrorlist.ChannelPublicGroupNaError):

                    return None, "У вас нет доступа к чату"
                except (ValueError):

                    return None, "Чат не существует"
                except errors.rpcerrorlist.FloodWaitError as e:
                    self.logger.info(
                        f"Аккаунт {self.me.phone} ожидает {e.seconds} секунд")
                    await asyncio.sleep(e.seconds + 5)
                    full = await self.client(functions.channels.GetFullChannelRequest(chat))

                full_channel = full.full_chat
                chat_id = full_channel.id
                test = await self.client.get_entity(chat_id)

                if (not test.megagroup and not test.gigagroup):
                    if full_channel.linked_chat_id:
                        chat_id = full_channel.linked_chat_id
                messages_count = (await self.client.get_messages(chat_id)).total
                message_current = 0
                while True:
                    try:
                        async for msg in self.client.iter_messages(chat_id, limit=messages_count):
                               sender = await msg.get_sender()
                                message_current += 1
                                if message_current == messages_count:
                                    self.logger.info(
                                        f"Аккаунт {self.me.phone} просмотрел все сообщения в чате {chat}")
                                    return True
                                if (sender and sender.__class__.__name__ == "User" and not sender.bot):
                                    if (not sender.stories_unavailable and not sender.stories_hidden and sender.stories_max_id and sender.username not in self.result and sender.username):
                                        identifier = sender.username
                                        try:
                                            stories = await self.client(GetPeerStoriesRequest(
                                                identifier,
                                            ))
                                        except FloodWaitError as e:
                                            self.logger.info(
                                                f"Аккаунт {self.me.phone} ожидает {e.seconds} секунд в чате {chat}")
                                            await asyncio.sleep(e.seconds)
                                            stories = await self.client(GetPeerStoriesRequest(
                                                identifier,
                                            ))
                                        if len(stories.stories.stories) == 0:
                                            self.logger.info(
                                                f"Аккаунт {self.me.phone} не имеет историй {identifier}")
                                            self.result[identifier] = 0
                                            await asyncio.sleep(rd(*looking))
                                            continue
                                        else:
                                            self.result[identifier] = 0
                                            user = stories.users[0]
                                            stories_to_view = stories.stories.stories[0]
                                            if not safe_mode:
                                                try:
                                                    stories = await self.client(ReadStoriesRequest(
                                                        user,
                                                        max_id=stories_to_view.id
                                                    ))
                                                except FloodWaitError as e:
                                                    self.logger.info(
                                                        f"Аккаунт {self.me.phone} ожидает {e.seconds} секунд в чате {chat}")
                                                    await asyncio.sleep(e.seconds)
                                                    stories = await self.client(ReadStoriesRequest(
                                                        user,
                                                        max_id=stories_to_view.id
                                                    ))
                                                await asyncio.sleep(rd(*looking))
                                                try:
                                                    await self.client(SendReactionRequest(
                                                        user,
                                                        stories_to_view.id,
                                                        reaction=ReactionEmoji(
                                                            emoticon=random.choice(
                                                                reactions),
                                                        )
                                                    ))
                                                except FloodWaitError as e:
                                                    self.logger.info(
                                                        f"Аккаунт {self.me.phone} ожидает {e.seconds} секунд")
                                                    await asyncio.sleep(e.seconds)
                                                    await self.client(SendReactionRequest(
                                                        user,
                                                        stories_to_view.id,
                                                        reaction=ReactionEmoji(
                                                            emoticon=random.choice(
                                                                reactions),
                                                        )
                                                    ))
                                                self.result[identifier] += 1
                                                self.logger.info(
                                                    f"Аккаунт {self.me.phone} посмотрел и поставил реакцию на историю {identifier}")
                                                await asyncio.sleep(rd(*reaction_flood))
                                            else:
                                                try:
                                                    stories = await self.client(ReadStoriesRequest(
                                                        user,
                                                        max_id=stories_to_view.id
                                                    ))
                                                except FloodWaitError as e:
                                                    self.logger.info(
                                                        f"Аккаунт {self.me.phone} ожидает {e.seconds} секунд в чате {chat}")
                                                    await asyncio.sleep(e.seconds)
                                                    stories = await self.client(ReadStoriesRequest(
                                                        user,
                                                        max_id=stories_to_view.id
                                                    ))
                                                self.result[identifier] += 1
                                                self.logger.info(
                                                    f"Аккаунт {self.me.phone} посмотрел историю {identifier}")
                                                await asyncio.sleep(rd(*looking))
                    except FloodWaitError as e:
                        self.logger.info(
                            f"Аккаунт {self.me.phone} ожидает {e.seconds} секунд в чате {chat}")
                        await asyncio.sleep(e.seconds)


        all_results = []
        tasks = []
        batch_size = 1
        for i in range(0, len(chats), batch_size):
            batch = chats[i:i + batch_size]
            for chat in batch:

                tasks.append(asyncio.create_task(masslook_chat(
                    chat, looking, reaction_flood, reactions)))
                await asyncio.sleep(rd(*wait))
            done, pending = await asyncio.wait(tasks, return_when=asyncio.ALL_COMPLETED)
            if not pending:
                results = [task.result() for task in done]
                print(results)
                all_results.extend(results)
            await asyncio.sleep(rd(*wait))
        if all(all_results):
            self.logger.info(
                "Не найдено историй в чатах для просмотра, ожидаю 2 часа!!..")
            await asyncio.sleep(7200)
            self.logger.info(
                "За 2 часа не было найдено историй, завершаю функцию..")
            return True
        else:
            return results

    async def MassTeging(self, wait, users, stories_url, photo, time, video= None):
        self.users = users
        time = int(time)
        try:
            photo = photo[0]
        except:
            pass
        stories_url = random.choice(stories_url)

        if not self.client.is_connected():
            status = await self.Check()
            if status is not True and "ограничен" not in status:
                return None, status

        stories = await self.client(functions.stories.GetAllStoriesRequest())
        ids = []
        for i in stories.peer_stories:
            if hasattr(i.peer, 'user_id'):
                if i.peer.user_id == self.me.id:
                    for story in i.stories:
                        ids.append(story.id)
            else:
                if i.peer.channel_id == self.me.id:
                    for story in i.stories:
                        ids.append(story.id)
        if ids != []:
            try:
                await self.client(functions.stories.DeleteStoriesRequest(
                    peer=self.me,
                    id=ids
                ))
            except:
                self.logger.error(
                    f'Клиент {self.me.phone} не смог удалить истории')

        users = group_list(users, 5)
        for i, user in enumerate(users):
            user = [f'@{us}' if "@" not in us else us for us in user]
            for us in user:
                sticker_sets = await self.client(GetAllStickersRequest(0))
                sticker_set = sticker_sets.sets[0]

                stickers = await self.client(GetStickerSetRequest(
                    stickerset=InputStickerSetID(
                        id=sticker_set.id, access_hash=sticker_set.access_hash
                    ),
                    hash=0
                ))
                sticker_ = random.choice(stickers.documents)
                try:
                    sticker_message = await self.client.send_file(us, file=sticker_)
                    self.logger.info(
                        "Отправка рандомного стикера пользователю {}".format(us))
                except Exception as e:
                    self.logger.error(
                        f"Аккаунт {self.me.phone} ограничен при попытке отправить стикер пользователю {us}")

                await asyncio.sleep(rd(*wait))

            string = str("\n".join(user))
            try:

                if self.me.premium and "https://" in stories_url:
                    vi = video[0]
                    if vi is not None:
                        res = await self.client(functions.stories.SendStoryRequest(
                            peer=self.me,
                            media_areas=[MediaAreaUrl(url=stories_url, coordinates=MediaAreaCoordinates(
                                50, 50, 82.12963104248, 31.041666030884, 0, 4.6277777353923),)],
                            media=types.InputMediaUploadedDocument(
                                file=await self.client.upload_file(vi),
                                mime_type="video/mp4",
                                attributes=[types.DocumentAttributeVideo(
                                    duration=38,
                                    w=720,
                                    h=1280,
                                    supports_streaming=True
                                )]

                            ),
                            privacy_rules=[types.InputPrivacyValueAllowAll()],
                            caption= string,
                            pinned=True,
                            noforwards=True,
                            period=3600*time,
                        ))
                        video.remove(vi)
                    else:
                        res = await self.client(functions.stories.SendStoryRequest(
                            peer=self.me,
                            media=types.InputMediaUploadedPhoto(
                                file=await self.client.upload_file(photo),

                            ),
                            media_areas=[MediaAreaUrl(url=stories_url, coordinates=MediaAreaCoordinates(
                                50, 50, 82.12963104248, 31.041666030884, 0, 4.6277777353923),)],
                            privacy_rules=[types.InputPrivacyValueAllowAll()],
                            caption = stories_url + "\n" + string,
                            pinned=True,
                            noforwards=True,
                            period=3600*time,
                        ))
                else:
                    vi = video[0]
                    if vi is not None:
                        res = await self.client(functions.stories.SendStoryRequest(
                            peer=self.me,
                            media=types.InputMediaUploadedDocument(
                                file=await self.client.upload_file(vi),
                                mime_type="video/mp4",
                                attributes=[types.DocumentAttributeVideo(
                                    duration=38,
                                    w=720,
                                    h=1280,
                                    supports_streaming=True
                                )]

                            ),
                            privacy_rules=[types.InputPrivacyValueAllowAll()],
                            caption=stories_url + "\n" + string,
                            pinned=True,
                            noforwards=True,
                            period=3600*time,
                        ))

                        video.remove(vi)
                    else:
                        res = await self.client(functions.stories.SendStoryRequest(
                            peer=self.me,
                            media=types.InputMediaUploadedPhoto(
                                file=await self.client.upload_file(photo),

                            ),

                            privacy_rules=[types.InputPrivacyValueAllowAll()],
                            caption = stories_url + "\n" + string,
                            pinned=True,
                            noforwards=True,
                            period=3600*time,
                        ))
                try:
                    await sticker_message.delete()
                except:
                    pass
                pinged = []
                for user_f in res.users:
                    if user_f.username is None:
                        continue
                    if "@" + user_f.username in user:
                        pinged.append(user_f.username)
                self.logger.info(
                    f"Клиент {self.me.phone} отметил пользователей {pinged} под историей")

                for us in pinged:
                    self.result[us] = True
                await asyncio.sleep(rd(*wait))
            except errors.rpcerrorlist.FloodError:
                self.logger.info(
                    f"Клиент {self.me.phone} не может создавать истории, превышен недельный лимит историй")
            except Exception as e:
                self.logger.error(e)
                return None, e

        return True

    async def AutoCommenting(self, channels, token, prompt):
        if not self.client.is_connected():
            status = await self.Check()
            if status is not True and "ограничен" not in status:
                return None, status

    async def _send_mention(self, chat_id, text=None, photo=None, video=None, send=None, wait=None):


        mentions = [{participant.username if participant.username else participant.id: f"<a href='tg://user?id={participant.id}'>ㅤ</a>"} async for participant in self.client.iter_participants(chat_id) if not participant.bot]
        usernames = [list(item.keys())[0] for item in mentions]
        ids = [list(item.values())[0] for item in mentions]
        mention_chunks = list(chunks(ids, 5))
        usernames_chunks = list(chunks(usernames, 5))

        for i, chunk in enumerate(mention_chunks):
            if text is not None:
                txt = text + f"\n{''.join(chunk)}"
            else:
                txt = f"\n{''.join(chunk)}"
            if photo is not None:
                photo = await self.client.upload_file(photo)
            result = await self.SendMessageMentions(chat_id, txt, video, photo, usernames_chunks[i])
            await asyncio.sleep(rd(*send))
            if type(result) is tuple:
                return result

        try:
            await self.client(functions.channels.LeaveChannelRequest(channel=chat_id))
        except Exception:
            pass

        return result

    async def SendMentions(self, chats, wait, send, text=None, photo=None, video=None):
        self.users = chats

        if not self.client.is_connected():
            status = await self.Check()
            if status is not True and "ограничен" not in status:
                return None, status
        for i, chat in enumerate(chats):
            try:
                await self.client(functions.channels.JoinChannelRequest(channel=chat))
            except (errors.ChannelsTooMuchError, errors.rpcerrorlist.ChannelPrivateError, errors.rpcerrorlist.InviteRequestSentError):
                return None, "У вас нет доступа к чату"
            except (errors.ChannelInvalidError, ValueError):
                return None, "Чат не существует"
            if i < len(chats) - 1:
                await asyncio.sleep(rd(*wait))
        tasks = []
        for i, chat_id in enumerate(chats):
            task = asyncio.create_task(self._send_mention(
                chat_id, text, photo, video, send, wait))
            tasks.append(task)

        results = await asyncio.gather(*tasks, return_exceptions=True)

        for i, result in enumerate(results):
            if isinstance(result, Exception):
                self.logger.error(f"Ошибка при отправке упоминания в чат {
                    chats[i]}: {result}")
            elif isinstance(result, tuple):
                return result
            elif result is True:
                # self.result[chats[i]] = result
                pass

        return True

    async def GetParticipants(self, chat, bio):

        try:
            result = {}
            async for user in self.client.iter_participants(chat):

                if user and not user.bot:
                    if hasattr(user.status, 'was_online'):
                        status = user.status.was_online
                    elif hasattr(user.status, 'expires') and not hasattr(user.status, 'was_online'):
                        status = "Онлайн"
                    elif user.status.__class__ is UserStatusEmpty:
                        status = "Давно"
                    elif user.status.__class__ is UserStatusLastWeek:
                        status = "Был на прошлой неделе"
                    elif user.status.__class__ is UserStatusLastMonth:
                        status = "Был в прошлом месяце"
                    elif user.status.__class__ is UserStatusRecently:
                        status = "Был недавно"
                    else:
                        status = "Был давно"
                    if (
                        not user.stories_unavailable and
                        not user.stories_hidden and
                        user.stories_max_id
                    ):
                        story = "Есть"
                    else:
                        story = "Нет"
                    if user.username:
                        if user.first_name:
                            first_name = user.first_name
                        else:
                            first_name = ""
                        if user.last_name:
                            last_name = user.last_name
                        else:
                            last_name = ""
                        try:
                            pol = predict(model, first_name + " " + last_name)
                        except Exception as e:
                            pol = "unknown"
                        result[user.id] = {
                            'username': user.username,
                            'phone': user.phone,
                            'status': status,
                            'premium': "Премиум" if user.premium else "Без премиума",
                            'story': story,
                            'pol': pol,
                            'bio': await GetBio(user.username) if bio else None,
                        }

        except (errors.rpcerrorlist.ChatAdminRequiredError, errors.rpcerrorlist.TimeoutError) as e:
            return None, "У вас нет доступа к участникам чата!"

        return result

    async def Parsing(self, chats, offset, max_accounts, wait, bio):
        if not self.client.is_connected():
            status = await self.Check()
            if status is not True and "ограничен" not in status:
                return None, status
        self.urls = chats

        async def parse_chat(chat):
            try:
                url = chat
                try:
                    if "joinchat" in url or "+" in url:
                        channel_id = url.split('/')[-1].replace('+', '')

                        await self.client(ImportChatInviteRequest(channel_id))
                        channel = await self.client.get_entity(url)
                        chat_id = channel.id
                    else:
                        channel = await self.client.get_entity(url)
                        await self.client(JoinChannelRequest(channel))
                        chat_id = channel.id
                    self.logger.info(
                        f'Клиент {self.me.phone} присоединился к {url}')
                except UserAlreadyParticipantError:
                    channel = await self.client.get_entity(url)
                    chat_id = channel.id
                except Exception as e:
                    self.logger.error(
                        f'Клиент {self.me.phone} не смог присоединиться к {url}: {e}')
                try:
                    full = await self.client(functions.channels.GetFullChannelRequest(chat))
                except (errors.rpcerrorlist.ChannelPrivateError,
                        errors.rpcerrorlist.TimeoutError,
                        errors.rpcerrorlist.ChannelPublicGroupNaError):

                    return None, "У вас нет доступа к чату"
                except (ValueError):

                    return None, "Чат не существует"

                full_channel = full.full_chat
                chat_id = full_channel.id
                test = await self.client.get_entity(chat_id)

                if (not test.megagroup and not test.gigagroup):
                    if full_channel.linked_chat_id:
                        chat_id = full_channel.linked_chat_id
                else:
                    if not full_channel.participants_hidden and full_channel.participants_count < 10000:
                        participants = await self.GetParticipants(chat_id, bio)
                        if type(participants) is tuple:
                            return participants
                        self.result[chat] = participants
                        return True

                messages_count = (await self.client.get_messages(chat_id)).total
                message_current = 0
                self.result[chat] = {}
                try:
                    async for msg in self.client.iter_messages(chat_id, add_offset=(messages_count // max_accounts) * offset, limit=messages_count):
                        sender = await msg.get_sender()

                        if (sender and sender.__class__.__name__ == "User" and not sender.bot and sender.id not in self.result[chat] and sender.username):

                            if hasattr(sender.status, 'was_online'):
                                status = sender.status.was_online.strftime(
                                    "%d.%m.%Y %H:%M:%S")
                            elif hasattr(sender.status, 'expires') and not hasattr(sender.status, 'was_online'):
                                status = "Онлайн"
                            else:
                                status = {
                                    UserStatusEmpty: "Давно",
                                    UserStatusLastWeek: "Был на прошлой неделе",
                                    UserStatusLastMonth: "Был в прошлом месяце",
                                    UserStatusRecently: "Был недавно"
                                }.get(type(sender.status), "Был давно")

                            story = "Есть" if (not sender.stories_unavailable and
                                               not sender.stories_hidden and
                                               sender.stories_max_id) else "Нет"

                            # chat_result[sender.id] = {
                            #     'username': sender.username,
                            #     'phone': sender.phone,
                            #     'status': status,
                            #     'premium': "Премиум" if sender.premium else "Без премиума",
                            #     'story': story
                            # }
                            if sender.first_name:
                                first_name = sender.first_name
                            else:
                                first_name = ""
                            if sender.last_name:
                                last_name = sender.last_name
                            else:
                                last_name = ""
                            try:
                                pol = predict(
                                    model, first_name + " " + last_name)
                            except Exception as e:
                                pol = "unknown"
                            self.result[chat][sender.id] = {
                                'username': sender.username,
                                'phone': sender.phone,
                                'status': status,
                                'premium': "Премиум" if sender.premium else "Без премиума",
                                'story': story,
                                'pol': pol,
                                'bio': await GetBio(sender.username) if bio else None,

                            }
                        message_current += 1
                except FloodWaitError as e:
                    await asyncio.sleep(e.seconds)
                # except asyncio.IncompleteReadError:
                #     await self.client.connect()
                # except ConnectionError:
                #     await self.client.connect()
                except Exception as e:
                    # await self.client.connect()
                    self.logger.error(e)
                # self.result[chat] = chat_result
            # except asyncio.IncompleteReadError:
            #     await self.client.connect()
            # except ConnectionError:
            #     await self.client.connect()
            except Exception as e:
                self.logger.error(e)
            return True

        tasks = []
        for idx, chat in enumerate(chats):
            tasks.append(asyncio.create_task(parse_chat(chat)))

            await asyncio.sleep(rd(*wait))

        results = await asyncio.gather(*tasks)
        if all(results):
            return True
        else:
            return results


    async def ParsingSearch(self, keywords, limit_search):
        if not self.client.is_connected():
            status = await self.Check()
            if status is not True and "ограничен" not in status:
                return None, status
        self.urls = keywords
        res = {}
        for keyword in keywords:
            result = await self.client(functions.contacts.SearchRequest(
                q=keyword,
                limit=limit_search
            ))
            chats = []
            for chat in result.chats:
                if chat.username:
                    chats.append(chat.username)
            res[keyword] = chats
        self.result = res
        return True


    async def Invite(self, chat, user):
        try:
            chats = await self.client.get_entity(chat)
            users = await self.client.get_entity(user)
            async for man in self.client.iter_participants(chat, search=user):
                if man.username and man.username.lower() == user.lower().replace('@', ''):
                    self.logger.info(f"Пользователь {user} уже в чате {chat}")
                    return True
            invite = await self.client(
                functions.channels.InviteToChannelRequest(
                    chats,
                    [users],
                )
            )
            # invite = await self.client(functions.messages.AddChatUserRequest(
            #     chat_id=chats.id,
            #     user_id=users.id,
            #     fwd_limit=42
            # ))
            if invite.missing_invitees:
                if invite.missing_invitees[0].premium_required_for_pm:
                    self.logger.info(
                        f"Пользователь {user} не может быть приглашен в {chat} (На аккаунте необходим ТГ премиум)")
                elif invite.missing_invitees[0].premium_would_allow_invite:
                    self.logger.info(
                        f"Пользователь {user} не может быть приглашен в {chat} (На аккаунте необходим ТГ премиум и нельзя пригласить из за настроек приватности.)")
                else:
                    self.logger.info(
                        f"Пользователь {user} не может быть приглашен в {chat} (Нельзя пригласить из за настроек приватности)")

            else:
                self.logger.info(f"Пользователь {user} приглашен в чат {chat}")
                return True
        except (errors.rpcerrorlist.ChatAdminRequiredError):
            return None, "Нужны админ-права"
        except (errors.rpcerrorlist.ChatWriteForbiddenError):
            return None, "У вас нет доступа к данному чату"
        except (errors.rpcerrorlist.UserNotMutualContactError):
            self.logger.error(
                f"Аккаунт {self.me.phone} получил флуд на добавление пользователей")
        except (
            errors.rpcerrorlist.InputUserDeactivatedError, errors.rpcerrorlist.PeerIdInvalidError,
            errors.rpcerrorlist.UserAlreadyParticipantError, errors.rpcerrorlist.UserIdInvalidError,
            errors.rpcerrorlist.UserPrivacyRestrictedError,
            errors.rpcerrorlist.PeerFloodError, errors.rpcerrorlist.UserKickedError,
            ValueError, errors.rpcerrorlist.UserChannelsTooMuchError, TypeError
        ) as e:
            self.logger.error(f"error: {str(e)}")

        except errors.rpcerrorlist.AuthKeyDuplicatedError:
            await self.client.connect()
        except asyncio.IncompleteReadError:
            await self.client.connect()
        except ConnectionError:
            await self.client.connect()
        except errors.rpcerrorlist.UsersTooMuchError:
            return None, "В чате уже набрано максимальное количество пользователей"
        except errors.rpcerrorlist.FloodWaitError:
            return None, f"Флуд-задержка превысила {self.client.flood_sleep_threshold} сек."
        except (errors.ChannelsTooMuchError, errors.rpcerrorlist.ChannelPrivateError, errors.rpcerrorlist.InviteRequestSentError):
            return None, "У вас нет доступа к чату"
        except (errors.ChannelInvalidError):
            await self.client.disconnect()
            return None, "Чат не существует"
        except Exception as e:
            self.logger.error(e)
        # else:
        #     return True

    async def Join(self, chat_id):
        try:
            await self.client(functions.channels.JoinChannelRequest(channel=chat_id))
        except (errors.ChannelsTooMuchError, errors.rpcerrorlist.ChannelPrivateError, errors.rpcerrorlist.InviteRequestSentError):
            return None, "У вас нет доступа к чату"
        except (errors.ChannelInvalidError, ValueError, UsernameInvalidError):
            return None, "Чат не существует"
        except errors.rpcerrorlist.FloodWaitError:
            return None, f"Флуд-задержка превысила {self.client.flood_sleep_threshold} сек."
        except Exception as e:
            return None, e
        return True

    async def Inviting(self, users, chat_id, wait, count=0, remaining= None):
        if not self.client.is_connected():
            status = await self.Check()
            if status is not True and "ограничен" not in status:
                return None, status
        count_true = 0
        wait = list(map(int, wait.split('-')))
        self.users = users
        try:
            for i, user in enumerate(users):
                await asyncio.sleep(0)
                result = await self.Invite(chat_id, user)
                self.result[user] = result
                if type(result) is tuple:
                    return result
                if result:
                    count_true += 1
                if i < len(users)-1:
                    await asyncio.sleep(rd(*wait))

        except ConnectionError:
            return None, "Подключение было оборвано"

        if remaining == []:
            return True
        while True:
            for i, user in enumerate(remaining):
                await asyncio.sleep(0)
                if count_true == count:
                    return True
                else:
                    result = await self.Invite(chat_id, user)
                    self.result[user] = result
                    if type(result) is tuple:
                        return result
                    if result:
                        count_true += 1
                        remaining.remove(user)
                    if i < len(users)-1:
                        await asyncio.sleep(rd(*wait))

            return True, f"Пользователи в базе закончились"

    async def ResetAuthorization(self):
        if not self.client.is_connected():
            status = await self.Check()
            if status is not True and "ограничен" not in status:

                return None, status
        count = 0
        try:
            auth = await self.client(functions.account.GetAuthorizationsRequest())
            for i in auth.authorizations:
                if not i.current:
                    try:
                        result = await self.client(functions.account.ResetAuthorizationRequest(i.hash))
                        count += 1
                        self.logger.info(
                            f"Авторизация была сброшена для аккаунта {self.me.phone}")
                    except errors.rpcerrorlist.FreshResetAuthorisationForbiddenError:
                        self.logger.error(
                            f"С момента входа в текущий сеанс прошло менее 24 часов. Нельзя сбросить авторизацию для аккаунта {self.me.phone}")
                        return None, "С момента входа в текущий сеанс прошло менее 24 часов"
                    except Exception as e:
                        self.logger.error(e)
                        return None, e
        except Exception as e:
            self.logger.error(e)
            return None, e
        if count == 0:
            self.logger.info(
                f"Аккаунт {self.me.phone} не имеет активных авторизаций, кроме текущей")
        try:
            await self.client.disconnect()
        except:
            pass
        return True

    async def New2fa(self, code, logger):
        if not self.client.is_connected():
            status = await self.Check()
            if status is not True and "ограничен" not in status:
                return None, status

        try:
            await self.client.edit_2fa(self.twofa, code)
            update_2fa(self.me.phone, code)
            logger.info(
                f"Клиент {self.me.phone} сменил 2fa с {self.twofa} на {code}")
        except Exception as e:
            self.logger.error(e)

        return True

    async def SetChannel(self, channel):
        if not self.client.is_connected():
            status = await self.Check()
            if status is not True and "ограничен" not in status:
                return None, status

        try:
            await self.client(functions.account.UpdatePersonalChannelRequest(channel=InputChannelEmpty()))
        except Exception as e:
            self.logger.error(e)
        await self.client(JoinChannelRequest(channel))
        channel_info = await self.client.get_entity(channel)
        full = await self.client(functions.channels.GetFullChannelRequest(channel))

        result = await self.client(functions.channels.CreateChannelRequest(
            title=channel_info.title,
            about=full.full_chat.about,
        ))
        id = str(random.randint(0, 9999))
        link = await self.client(functions.channels.UpdateUsernameRequest(channel=result.chats[0], username=channel_info.username+id))

        self.logger.info(
            f"Клиент {self.me.phone} создал канал  https://t.me/{channel_info.username+id}")
        res = await self.client.download_profile_photo(channel_info, "profile.jpg")
        if res:
            await self.client(functions.channels.EditPhotoRequest(
                channel=result.chats[0],
                photo=await self.client.upload_file('profile.jpg'),
            ))
        count = 0
        async for post in self.client.iter_messages(channel_info, reverse=True):

            try:

                await self.client.forward_messages(entity=result.chats[0], from_peer=channel_info, messages=[post.id], drop_author=True)
                self.logger.info(
                    f"Клиент {self.me.phone} переслал пост в канал https://t.me/{channel_info.username+id}")
                count += 1
            except Exception as e:
                pass
        res = await self.client(functions.channels.GetAdminedPublicChannelsRequest(for_personal=True))
        try:
            res = await self.client(functions.account.UpdatePersonalChannelRequest(channel=channel_info.username+id))
            me = await self.client.get_me()
        except Exception as e:
            self.logger.error(e)

        return True

    async def CallPhone(self, wait, users, video=False):
        self.users = users
        if not self.client.is_connected():
            status = await self.Check()
            if status is not True and "ограничен" not in status:
                return None, status
        # g_a_hash = await self.client(functions.messages.GetDhConfigRequest(
        #         version=0,
        #         random_length=256
        #     ))
        # g_a_hash = generate_g_a_hash(g_a_hash.p, g_a_hash.g)
        for user in users:
            g_a_hash = await self.client(functions.messages.GetDhConfigRequest(
                version=0,
                random_length=256
            ))
            g_a_hash = generate_g_a_hash(g_a_hash.p, g_a_hash.g)
            try:

                result = await self.client(functions.phone.RequestCallRequest(
                    user_id=f"@{user}" if '@' not in user else user,
                    g_a_hash=g_a_hash,
                    protocol=types.PhoneCallProtocol(
                        min_layer=65,
                        max_layer=65,
                        udp_p2p=True,
                        library_versions=["2.4.4", "2.7"]
                    ),
                    video=video
                ))
                res = await self.client(functions.phone.DiscardCallRequest(
                    peer=types.InputPhoneCall(
                        id=result.phone_call.id,
                        access_hash=result.phone_call.access_hash,
                    ),
                    duration=0,
                    reason=types.PhoneCallDiscardReasonMissed(),
                    connection_id=0,
                    video=video
                ))
                self.logger.info(
                    f"Звонок с {self.me.phone} на {user} был отправлен и завершен")
                self.result[user] = True
            except errors.rpcerrorlist.UserPrivacyRestrictedError:
                self.logger.info(
                    f"Звонок с {self.me.phone} на {user} невозможен из-за настроек приватности пользователя")
            except errors.rpcerrorlist.UserIsBlockedError:
                self.logger.info(
                    f"Звонок с {self.me.phone} на {user} невозможен, пользователь заблокировал аккаунт")
            except Exception as e:
                self.logger.error(e)
            await asyncio.sleep(rd(*wait))
        return True

    async def WireTapping(self, wait, groups, keywords, admin):
        self.users = groups
        if not self.client.is_connected():
            status = await self.Check()
            if status is not True and "ограничен" not in status:
                return None, status
        if "joinchat" in admin or "+" in admin:
               channel_id = admin.split('/')[-1].replace('+', '')
                try:
                    await self.client(ImportChatInviteRequest(channel_id))
                except:
                    pass
                channel = await self.client.get_entity(admin)
                chat_id = channel.id
        else:
            channel = await self.client.get_entity(admin)
            await self.client(JoinChannelRequest(channel))
            chat_id = channel.id
        for group in groups:
            channel = await self.client.get_input_entity(group)
            await self.client(JoinChannelRequest(channel))
            if hasattr(channel, 'channel_id'):
                id = channel.channel_id
            elif hasattr(channel, 'id'):
                id = channel.id
            else:
                id = channel
            self.client.add_event_handler(lambda event: message_handler_wiretapping(
                event, wait, keywords, chat_id, group), events.NewMessage(chats=[id]))
        while True:
            await asyncio.sleep(1)


    async def Reaction1(self, urls, reactions, wait):
        if not self.client.is_connected():
            status = await self.Check()
            if status is not True and "ограничен" not in status:
                return None, status

        self.urls = urls
        for url in urls:
            channel = await self.client.get_entity(url)
            await self.client(JoinChannelRequest(channel))
            self.client.add_event_handler(lambda event: reaction1(
                event, channel, reactions, url, wait, self.logger, self.me.phone), events.NewMessage(chats=[channel.id]))
        while True:
            await asyncio.sleep(1)

    async def Reaction2(self, urls, reactions, time):
        self.urls = urls
        if not self.client.is_connected():
            status = await self.Check()
            if status is not True and "ограничен" not in status:
                return None, status
        for url in urls:
            channel = await self.client.get_entity(url)
            await self.client(JoinChannelRequest(channel))
            self.client.add_event_handler(lambda event: reaction2(
                event, channel, reactions, datetime.now(tz=timezone.utc), time, url), events.NewMessage(chats=[channel.id]))
        while True:
            await asyncio.sleep(1)

    async def Reaction3(self, urls, reactions, wait):
        self.urls = urls
        if not self.client.is_connected():
            status = await self.Check()
            if status is not True and "ограничен" not in status:
                return None, status
        for url in urls:

            channel = await self.client.get_entity(url)
            await self.client(JoinChannelRequest(channel))
            self.client.add_event_handler(lambda event: reaction3(
                event, channel, reactions, wait, url, self.logger), events.NewMessage(chats=[channel.id]))
        while True:
            callback = self.client.list_event_handlers()
            if len(callback) == 0:
                return True
            await asyncio.sleep(1)

    async def SuggestPhoto(self, wait, users, suggest_photo, text=None, photo=None, video=None, message_id=None, channel_id=None):
        self.users = users
        if not self.client.is_connected():
            status = await self.Check()
            if status is not True and "ограничен" not in status:
                return None, status
        suggest_photo = suggest_photo[0]
        for user in users:
            while True:
                try:
                    if channel_id or message_id:
                        await self.client.forward_messages(entity=user, from_peer=channel_id, messages=[message_id])
                        self.logger.info(
                            f"Пересылка сообщения пользователю {user}")
                    elif photo is not None:
                        await self.client.send_file(user, file=photo, caption=text, parse_mode='html')
                    elif video is not None:
                        await self.client.send_file(user, file=video, caption=text, parse_mode='html', supports_streaming=True)
                    else:
                        await self.client.send_message(user, text, parse_mode='html')
                        self.logger.info(
                            f"Отправка сообщения пользователю {user}")

                    await self.client(functions.photos.UploadContactProfilePhotoRequest(
                        user_id=f"@{user}" if '@' not in user else user,
                        suggest=True,
                        file=await self.client.upload_file(suggest_photo),
                    ))
                    self.result[user] = True
                    await asyncio.sleep(rd(*wait))

                    break
                except errors.rpcerrorlist.SlowModeWaitError as e:
                    await asyncio.sleep(e.seconds)
                    continue
                except (
                    errors.rpcerrorlist.UserIsBlockedError, errors.rpcerrorlist.ChatAdminRequiredError,
                    errors.rpcerrorlist.ChannelPrivateError, errors.rpcerrorlist.ChatWriteForbiddenError,
                    errors.rpcerrorlist.InputUserDeactivatedError, errors.rpcerrorlist.TimeoutError,
                    ValueError, errors.rpcerrorlist.UsernameInvalidError,
                    errors.rpcerrorlist.ChatIdInvalidError, errors.rpcerrorlist.ChatRestrictedError, errors.rpcerrorlist.EntityBoundsInvalidError,
                    errors.rpcerrorlist.EntityMentionUserInvalidError, errors.rpcerrorlist.PeerIdInvalidError, errors.rpcerrorlist.EntityBoundsInvalidError,
                    OperationalError
                ) as e:
                    self.logger.error(f"error: {str(e)}")
                    return None, str(e)
                except errors.rpcerrorlist.PeerFloodError:
                    self.logger.error(f"Аккаунт {self.me.phone} ограничен")
                    return None, "Аккаунт ограничен"

                except errors.rpcerrorlist.AuthKeyDuplicatedError:
                    return None, "Авторизован с другого IP-адреса"
                except errors.rpcerrorlist.FloodWaitError:
                    return None, f"Флуд-задержка превысила {self.client.flood_sleep_threshold} сек."
                except (ConnectionError):
                    return None, "Подключение было оборвано"
                except Exception as e:
                    self.logger.error(f"error: {str(e)}")
        os.remove(suggest_photo)
        return True

    async def AddContacts(self, wait, users):
        self.users = users
        if not self.client.is_connected():
            status = await self.Check()
            if status is not True and "ограничен" not in status:
                return None, status
        for user in users:
            while True:
                try:
                    res = await self.client(functions.contacts.AddContactRequest(
                        id=user,
                        first_name="User",
                        last_name="",
                        phone="",
                        add_phone_privacy_exception=True
                    ))
                    self.logger.info(
                        f"Добавление контакта пользователя {user}")
                    self.result[user] = True
                    await asyncio.sleep(rd(*wait))
                    break
                except Exception as e:
                    self.logger.error(f"error: {str(e)}")
                    await asyncio.sleep(rd(*wait))

        return True
